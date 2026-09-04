"""The KB workbook (c18/c19): export for review, import as proposals.

THE accept, from v1: export, change nothing, import — and the engine
concludes nothing happened. If that breaks, every import starts with
noise a steward has to read past, which is how review becomes
rubber-stamping.

Also decides, explicitly, the seven failure modes v1 left unhandled
(B40/D22). Each has a test here, because "we thought about it" and "it
does the right thing" are different claims.
"""

import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from engine.flywheel.proposals import ProposalStore
from engine.kb.store import KBStore
from engine.kb.xlsx import (LIST_SEP, READ_ONLY, SHEET, WorkbookError,
                            columns, export_cards, plan_import, read_workbook,
                            submit_import)

AT = "2026-08-05T00:00:00Z"
PROV = {"source_pursuit": "pur_x", "source_client": "Fixture County",
        "date": "2026-01-01", "ingested_by": "ingestion_agent"}


@pytest.fixture
def store(tmp_path):
    store = KBStore(tmp_path / "kb")
    store.write_card(
        {"kb_id": "kb_alpha0001", "layer": "corpus",
         "doc_kind": "section_exemplar", "title": "Data Migration Approach",
         "summary": "Seven mock conversions.", "owner": "Delivery Lead",
         "industries": ["healthcare", "public sector"],
         "canonical_block": False},
        "Body one.", PROV, {})
    store.write_card(
        {"kb_id": "kb_beta00001", "layer": "fact_sheet", "doc_kind": "fact",
         "title": "SOC 2", "summary": "Current attestation.",
         "claim_tier_max": 1, "verified_date": "2026-06-15",
         "owner": "Compliance Lead", "review_due": "2027-06-15"},
        "Body two.", PROV, {})
    return store


def _sheet(path):
    return load_workbook(path)[SHEET]


def _edit_cell(path, kb_id, column, value):
    book = load_workbook(path)
    sheet = book[SHEET]
    names = [c.value for c in sheet[1]]
    col = names.index(column) + 1
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == kb_id:
            sheet.cell(row=row, column=col, value=value)
            break
    book.save(path)
    return path


# ------------------------------------------------------------ THE accept

def test_an_unedited_round_trip_changes_nothing(store, tmp_path):
    path = export_cards(store, tmp_path / "kb.xlsx")
    plan = plan_import(store, path)
    assert plan["status"] == "ok"
    assert plan["changes"] == []
    assert plan["errors"] == []
    assert plan["unchanged"] == 2


def test_the_sheet_carries_enough_to_see_every_card(store, tmp_path):
    path = export_cards(store, tmp_path / "kb.xlsx")
    sheet = _sheet(path)
    assert sheet.max_row == 3          # header + two cards
    assert [c.value for c in sheet[1]] == [n for n, _ in columns()]


def test_columns_are_derived_from_the_schema_not_hand_listed():
    """v1's lesson: a hand-kept column list is the registry that drifts."""
    schema = json.loads(
        (Path(__file__).resolve().parents[2] / "schemas"
         / "kb-card.schema.json").read_text(encoding="utf-8"))
    declared = set(schema["properties"]) - {"anonymization", "provenance",
                                            "content_path",
                                            # WP13 (B59): structured /
                                            # navigation-only fields a flat
                                            # sheet cannot carry honestly.
                                            "identity", "chunk_span",
                                            "doc_path",
                                            # P26c: structured records
                                            "lessons", "deprecated"}
    assert {name for name, _ in columns()} == declared


def test_cells_are_typed_not_string_coerced(store, tmp_path):
    """P9's inline export str()-coerced everything, so a boolean arrived
    as the text 'False' and a number as a quoted string."""
    path = export_cards(store, tmp_path / "kb.xlsx")
    sheet = _sheet(path)
    names = [c.value for c in sheet[1]]
    tier_col = names.index("claim_tier_max") + 1
    values = [sheet.cell(row=r, column=tier_col).value
              for r in range(2, sheet.max_row + 1)]
    assert 1 in values, "an integer must survive as an integer"


def test_export_is_byte_stable_for_an_unchanged_pack(store, tmp_path):
    """openpyxl stamps dcterms:modified from the wall clock inside
    save(); pinning it keeps an unchanged pack producing an unchanged
    file (v1's intermittent-flake lesson)."""
    first = export_cards(store, tmp_path / "one.xlsx").read_bytes()
    second = export_cards(store, tmp_path / "two.xlsx").read_bytes()
    assert first == second


# ------------------------------------------ v1's seven unhandled modes

def test_mode_1_merged_cells_refuse_by_name(store, tmp_path):
    """openpyxl returns None for every cell of a merge but the first, so
    v1 reported a misleading 'required field empty'."""
    path = export_cards(store, tmp_path / "kb.xlsx")
    book = load_workbook(path)
    book[SHEET].merge_cells("C2:D2")
    book.save(path)
    with pytest.raises(WorkbookError) as caught:
        read_workbook(path)
    assert "C2:D2" in str(caught.value)
    assert "unmerge" in str(caught.value)


def test_mode_2_a_formula_is_a_cell_addressed_error(store, tmp_path):
    """v1 read formulas with data_only=True, so a workbook saved by a
    non-Excel tool had no cached value and the cell looked empty."""
    path = _edit_cell(export_cards(store, tmp_path / "kb.xlsx"),
                      "kb_alpha0001", "summary", "=CONCAT(A1,B1)")
    plan = plan_import(store, path)
    assert plan["status"] == "refused"
    formula = next(e for e in plan["errors"] if e["code"] == "formula_cell")
    assert "paste values" in formula["message"]
    assert formula["cell"].startswith(f"{SHEET}!")


def test_mode_3_header_matching_is_case_insensitive(store, tmp_path):
    """v1 matched case-sensitively, so 'ID' instead of 'id' refused the
    whole file."""
    path = export_cards(store, tmp_path / "kb.xlsx")
    book = load_workbook(path)
    sheet = book[SHEET]
    for cell in sheet[1]:
        cell.value = str(cell.value).upper()
    book.save(path)
    rows, errors, _warnings = read_workbook(path)
    assert rows and not errors


def test_mode_4_duplicate_columns_refuse_rather_than_first_wins(store,
                                                               tmp_path):
    """v1 silently kept the first and ignored the second."""
    path = export_cards(store, tmp_path / "kb.xlsx")
    book = load_workbook(path)
    sheet = book[SHEET]
    sheet.cell(row=1, column=sheet.max_column + 1, value="summary")
    book.save(path)
    with pytest.raises(WorkbookError) as caught:
        read_workbook(path)
    assert "duplicate column" in str(caught.value)


def test_mode_5_a_list_value_containing_the_separator_refuses(tmp_path):
    """v1's real data-loss bug: lists joined on ',' and split back on
    ',', so a tag containing a comma became two tags, silently and
    permanently. Refusing loudly is the fix."""
    store = KBStore(tmp_path / "kb")
    store.write_card(
        {"kb_id": "kb_gamma0001", "layer": "corpus", "summary": "S",
         "industries": [f"health{LIST_SEP.strip()}care"]},
        "Body.", PROV, {})
    with pytest.raises(WorkbookError) as caught:
        export_cards(store, tmp_path / "kb.xlsx")
    assert "round-trip" in str(caught.value)


def test_mode_6_dates_normalise_from_either_form(store, tmp_path):
    """Excel retypes an ISO date into a datetime; both must land as the
    same string, or every export/import cycle shows a phantom edit."""
    from datetime import datetime

    path = _edit_cell(export_cards(store, tmp_path / "kb.xlsx"),
                      "kb_beta00001", "verified_date",
                      datetime(2026, 6, 15))
    plan = plan_import(store, path)
    assert plan["status"] == "ok"
    assert plan["changes"] == [], "a retyped date is not an edit"


def test_mode_7_a_row_cap_refuses_a_runaway_sheet(store, tmp_path):
    from engine.kb import xlsx

    path = export_cards(store, tmp_path / "kb.xlsx")
    book = load_workbook(path)
    sheet = book[SHEET]
    for _ in range(12):
        sheet.append(["kb_filler"] + [None] * (sheet.max_column - 1))
    book.save(path)
    original = xlsx.MAX_ROWS
    try:
        xlsx.MAX_ROWS = 5
        with pytest.raises(WorkbookError) as caught:
            read_workbook(path)
        assert "row cap" in str(caught.value) or "exceeds" in str(caught.value)
    finally:
        xlsx.MAX_ROWS = original


# ------------------------------------------------ the v1 KEEP behaviours

def test_a_banner_row_above_the_header_is_tolerated(store, tmp_path):
    path = export_cards(store, tmp_path / "kb.xlsx")
    book = load_workbook(path)
    sheet = book[SHEET]
    sheet.insert_rows(1, 2)
    sheet["A1"] = "KNOWLEDGE BASE REVIEW — return by Friday"
    book.save(path)
    rows, errors, _ = read_workbook(path)
    assert len(rows) == 2 and not errors


def test_a_missing_known_column_refuses_before_reading_rows(store, tmp_path):
    path = export_cards(store, tmp_path / "kb.xlsx")
    book = load_workbook(path)
    book[SHEET].delete_cols(4)
    book.save(path)
    with pytest.raises(WorkbookError) as caught:
        read_workbook(path)
    assert "missing column" in str(caught.value)


def test_an_unknown_column_warns_once_and_is_ignored(store, tmp_path):
    """Someone adding a scratch 'notes' column has done nothing wrong —
    but a MISSPELLED header would otherwise vanish into that tolerance,
    so it is reported."""
    path = export_cards(store, tmp_path / "kb.xlsx")
    book = load_workbook(path)
    sheet = book[SHEET]
    sheet.cell(row=1, column=sheet.max_column + 1, value="reviewer notes")
    book.save(path)
    _rows, errors, warnings = read_workbook(path)
    assert not errors
    assert len(warnings) == 1 and "reviewer notes" in warnings[0]


def test_blank_rows_are_skipped(store, tmp_path):
    path = export_cards(store, tmp_path / "kb.xlsx")
    book = load_workbook(path)
    book[SHEET].append([None] * 5)
    book.save(path)
    rows, errors, _ = read_workbook(path)
    assert len(rows) == 2 and not errors


def test_a_wrong_workbook_is_refused_by_name(tmp_path):
    book = Workbook()
    book.active.title = "Sheet1"
    book.active["A1"] = "not a KB export"
    path = tmp_path / "other.xlsx"
    book.save(path)
    with pytest.raises(WorkbookError) as caught:
        read_workbook(path)
    assert "KB export" in str(caught.value)


def test_governance_columns_refuse_a_cell_edit(store, tmp_path):
    """'Approving' or reclassifying is a governance decision, not a cell
    edit — the sheet shows it and will not import a change."""
    path = _edit_cell(export_cards(store, tmp_path / "kb.xlsx"),
                      "kb_alpha0001", "canonical_block", True)
    plan = plan_import(store, path)
    assert plan["status"] == "refused"
    locked = next(e for e in plan["errors"] if e["code"] == "locked_field")
    assert "governance decision" in locked["message"]
    assert "canonical_block" in READ_ONLY


def test_a_derived_signal_cannot_be_typed_over(store, tmp_path):
    """edit_survival is measured, not asserted: a typed-over value would
    be a measurement nobody took."""
    path = _edit_cell(export_cards(store, tmp_path / "kb.xlsx"),
                      "kb_alpha0001", "edit_survival", 0.99)
    plan = plan_import(store, path)
    assert plan["status"] == "refused"
    assert any(e["code"] == "locked_field" for e in plan["errors"])


def test_a_new_kb_id_is_refused_with_the_reason(store, tmp_path):
    path = export_cards(store, tmp_path / "kb.xlsx")
    book = load_workbook(path)
    sheet = book[SHEET]
    sheet.append(["kb_brandnew1"] + [None] * (sheet.max_column - 1))
    book.save(path)
    plan = plan_import(store, path)
    assert plan["status"] == "refused"
    unknown = next(e for e in plan["errors"] if e["code"] == "unknown_kb_id")
    assert "provenance" in unknown["message"]


def test_duplicate_rows_name_the_other_row(store, tmp_path):
    path = export_cards(store, tmp_path / "kb.xlsx")
    book = load_workbook(path)
    sheet = book[SHEET]
    sheet.append([sheet.cell(row=2, column=c).value
                  for c in range(1, sheet.max_column + 1)])
    book.save(path)
    plan = plan_import(store, path)
    duplicate = next(e for e in plan["errors"]
                     if e["code"] == "duplicate_kb_id")
    assert "row 2" in duplicate["message"]


# ------------------------------------------------- all-or-nothing + S4

def test_one_error_stops_every_proposal(store, tmp_path):
    """A partial import leaves a steward guessing which of their forty
    edits landed, and the pack in a state nobody chose."""
    path = export_cards(store, tmp_path / "kb.xlsx")
    _edit_cell(path, "kb_alpha0001", "summary", "A good edit.")
    _edit_cell(path, "kb_beta00001", "claim_tier_max", "not a number")

    result = submit_import(store, path, operator="steward", at=AT)
    assert result["status"] == "refused"
    assert result["proposals"] == []
    assert ProposalStore(store.root).list() == []
    assert any(e["code"] == "not_a_number" for e in result["errors"])


def test_every_error_is_listed_never_a_sample(store, tmp_path):
    path = export_cards(store, tmp_path / "kb.xlsx")
    _edit_cell(path, "kb_alpha0001", "claim_tier_max", "nope")
    _edit_cell(path, "kb_beta00001", "claim_tier_max", "also nope")
    plan = plan_import(store, path)
    assert len([e for e in plan["errors"] if e["code"] == "not_a_number"]) == 2


def test_a_clean_import_opens_proposals_and_writes_no_card(store, tmp_path):
    """S4: the sheet proposes; a human disposes."""
    path = _edit_cell(export_cards(store, tmp_path / "kb.xlsx"),
                      "kb_alpha0001", "summary", "Seven mock conversions, "
                                                 "reconciled to the penny.")
    before = store.read_card("kb_alpha0001")

    result = submit_import(store, path, operator="steward", at=AT)
    assert result["status"] == "ok"
    assert len(result["proposals"]) == 1
    assert "nothing happened to them" in result["message"]
    assert store.read_card("kb_alpha0001") == before, "no card was written"

    proposal = ProposalStore(store.root).list()[0]
    assert proposal["kb_id"] == "kb_alpha0001"
    assert proposal["source"]["door"] == "xlsx_import"
    assert proposal["source"]["operator"] == "steward"
    assert proposal["diff"]["summary"]["before"] == "Seven mock conversions."


def test_restricted_provenance_never_reaches_the_sheet(store, tmp_path):
    """D22's rule, carried into the extracted module: card fronts carry
    no provenance by design, and the export must not add any."""
    path = export_cards(store, tmp_path / "kb.xlsx")
    text = path.read_bytes()
    assert b"Fixture County" not in text
    assert b"source_client" not in text
    assert "provenance" not in {name for name, _ in columns()}
