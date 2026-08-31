"""Structure + voice component suites (c6).

Both lanes are deterministic, so the interesting proofs are not "does it
pass" but "can it fail" and "does it stay put": a mutation must move each
number, and the committed voice record must keep describing the live
spec (the drift lock that makes "the owner blesses every wording change"
survive the owner not being in the room).
"""

import json
from pathlib import Path

import pytest

from engine.evals.structure import (ADVERSARIAL_GOLDENS, TWIN_GOLDENS,
                                    build_adversarial, evaluate_structure_set)
from engine.evals.voice import (CASES_PATH, RECORDED_PATH, evaluate_voice_set,
                                spec_fingerprint)


# ---------------------------------------------------------------- structure

@pytest.fixture(scope="module")
def structure_report(tmp_path_factory):
    return evaluate_structure_set(tmp_path_factory.mktemp("adv"))


def test_structure_parses_every_case_exactly(structure_report):
    assert structure_report["n_cases"] == len(TWIN_GOLDENS) + len(
        ADVERSARIAL_GOLDENS)
    assert structure_report["failures"] == []
    assert structure_report["exact_match"] == 1.0


def test_adversarial_workbooks_really_carry_their_trait(tmp_path):
    """The absence twin: an adversarial case that isn't actually
    adversarial proves nothing. Read the generated files back and prove
    the banner rows, the merge, and the spacer rows are really there."""
    from openpyxl import load_workbook

    built = build_adversarial(tmp_path)

    banner = load_workbook(built["adv_banner_rows"]).active
    assert banner["A1"].value.startswith("REQUEST FOR PROPOSAL")
    assert banner["A3"].value == "Ref", "the real header is not row 1"

    merged = load_workbook(built["adv_merged_prompt"]).active
    assert "B2:C2" in {str(r) for r in merged.merged_cells.ranges}

    spacers = load_workbook(built["adv_blank_spacers"]).active
    assert all(c.value is None for c in spacers[3]), "row 3 is a spacer"
    assert spacers["A4"].value == "4.2", "answers resume after the spacer"


def test_structure_suite_can_fail(tmp_path):
    """A suite that cannot fail proves nothing: move one golden and the
    lane must report the mismatch rather than round it away."""
    import engine.evals.structure as mod

    original = dict(mod.ADVERSARIAL_GOLDENS)
    try:
        mod.ADVERSARIAL_GOLDENS["adv_banner_rows"] = 99
        report = evaluate_structure_set(tmp_path)
    finally:
        mod.ADVERSARIAL_GOLDENS.clear()
        mod.ADVERSARIAL_GOLDENS.update(original)
    assert report["exact_match"] < 1.0
    assert any("adv_banner_rows" in f for f in report["failures"])


# -------------------------------------------------------------------- voice

@pytest.fixture(scope="module")
def voice_report():
    return evaluate_voice_set()


def test_voice_catches_every_approved_term_and_no_benign(voice_report):
    assert voice_report["recall"] == 1.0
    assert voice_report["misses"] == []
    assert voice_report["false_positives"] == []
    assert voice_report["benign_total"] == 5


def test_every_planted_term_is_really_in_the_committed_list():
    """Fixture integrity: a case planting a term the spec never listed
    would be unfalsifiable — it could only ever miss."""
    from engine.evals.cases import load_cases
    from engine.evals.voice import VOICE_SPEC
    from engine.validation.voice import prohibited_terms

    terms = set(prohibited_terms(VOICE_SPEC))
    planted = {label for case in load_cases(CASES_PATH)
               for label in case["expected"].get("labels", [])}
    assert planted, "the must-flag half must name its terms"
    assert planted <= terms


def test_the_header_row_is_not_an_enforced_term():
    """P10-F11 regression: `\\s*` before a lookahead backtracks, so the
    old row regex admitted the table header and 'prohibited' became an
    enforced term nobody approved. The benign case voice_ben_004 is the
    live proof — it would have flagged before the fix."""
    from engine.evals.voice import VOICE_SPEC
    from engine.validation.voice import prohibited_terms, voice_findings

    terms = prohibited_terms(VOICE_SPEC)
    assert "prohibited" not in terms
    assert voice_findings(
        "s1", "Access to prohibited transaction combinations is blocked.",
        terms) == []


def test_voice_suite_can_fail_when_the_list_loses_a_term(tmp_path):
    """The mutation proof: drop a term from a copy of the spec and recall
    falls. This is also the promotion gate in miniature — a spec edit
    that loses detection cannot pass silently."""
    from engine.evals.voice import VOICE_SPEC

    spec = VOICE_SPEC.read_text(encoding="utf-8")
    weakened = tmp_path / "voice-spec.md"
    weakened.write_text(spec.replace("| utilize | use |\n", ""),
                        encoding="utf-8")

    import engine.evals.voice as mod
    original = mod.VOICE_SPEC
    try:
        mod.VOICE_SPEC = weakened
        report = evaluate_voice_set()
    finally:
        mod.VOICE_SPEC = original
    assert report["recall"] < 1.0
    assert "voice_pro_002" in report["misses"]


def test_committed_voice_record_still_describes_the_live_spec(voice_report):
    """The drift lock (injection-suite pattern): editing config/voice-spec.md
    fails here until `python -c "from engine.evals.voice import
    write_recorded; write_recorded()"` re-derives the record CONSCIOUSLY."""
    committed = json.loads(RECORDED_PATH.read_text(encoding="utf-8"))
    assert committed == voice_report
    assert committed["spec_fingerprint"] == spec_fingerprint()


def test_both_lanes_report_through_the_release_record():
    from engine.evals.run import structure_lane, voice_lane

    for lane in (structure_lane, voice_lane):
        entry = lane()
        assert entry["basis"] == "deterministic"
        assert entry["blocking"] is True
        assert entry["bar"] and entry["measures"]
