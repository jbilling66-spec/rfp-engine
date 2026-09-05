"""P1-19 (P26b-3, the owner's call B119 §1a): the xlsx write-back hands
back the buyer's workbook byte-for-byte plus the answers.

openpyxl's save blanks every cached formula value and rewrites every
part; the writer is now a zip-level patch of the answered cells inside
their sheet XML, everything else copied verbatim, and `assert_roundtrip`
proves it after every write. The twin carries the part classes that
used to be at risk — a chart and its drawing, a cell comment and its
VML, a data validation, and a formula with a spliced cached value.
"""

import hashlib
import json
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from engine.assembly.writeback import run_writeback
from engine.assembly.xlsx_patch import (assert_roundtrip, sheet_parts,
                                        split_coord, write_cells)
from engine.contracts import ContractError, validate
from engine.llm import effective_config
from engine.runlog import RunLogger
from engine.version import engine_version
from engine.workspace import PursuitDir
from tests.helpers import plant_freeze

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"
TWIN = FIXTURES / "writeback-twin.xlsx"
AT = "2026-09-04T12:00:00Z"
FIRM = {"name": "Fixture Advisory LLP", "company": "Fixture Advisory LLP",
        "configured": True}
PROSE = "Founded in 2001, employee-owned, ERP delivery is the practice."
YELLOW = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00",
                     fill_type="solid")


def _members(path: Path) -> dict[str, bytes]:
    with zipfile.ZipFile(path) as zf:
        return {name: zf.read(name) for name in zf.namelist()}


def _unit_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "hello"          # a shared string
    ws["B1"] = "styled"         # keeps its style through the patch
    ws["B1"].fill = YELLOW
    ws["D1"] = "=1+1"           # a formula the answer overwrites
    ws["A2"] = "bystander"
    ws["E2"] = "=A2"            # a formula the answer leaves alone
    wb.save(path)
    return path


def test_split_coord_and_sheet_parts(tmp_path):
    assert split_coord("A1") == (1, 1)
    assert split_coord("AB12") == (28, 12)
    with pytest.raises(ContractError, match="not a cell coordinate"):
        split_coord("1A")
    with zipfile.ZipFile(TWIN) as zf:
        parts = sheet_parts(zf)
    assert parts == {"Instructions": "xl/worksheets/sheet1.xml",
                     "1. Questions": "xl/worksheets/sheet2.xml",
                     "Pricing": "xl/worksheets/sheet3.xml"}


def test_the_patch_writes_inline_strings_and_keeps_everything_else(tmp_path):
    source = _unit_workbook(tmp_path / "src.xlsx")
    output = tmp_path / "out.xlsx"
    writes = {("S", "A1"): "x", ("S", "B1"): "y", ("S", "C5"): "z",
              ("S", "D1"): "w"}
    result = write_cells(source, output, writes, firm=FIRM, at=AT)
    assert result["patched_parts"] == ["docProps/core.xml",
                                       "xl/worksheets/sheet1.xml"]
    assert result["formula_cells_overwritten"] == ["D1"]
    ws = load_workbook(output)["S"]
    assert (ws["A1"].value, ws["B1"].value, ws["C5"].value,
            ws["D1"].value) == ("x", "y", "z", "w")
    assert ws["B1"].fill.start_color.rgb == "FFFFFF00", "style kept"
    assert ws["A2"].value == "bystander" and ws["E2"].value == "=A2"
    src, out = _members(source), _members(output)
    assert list(src) == list(out)
    for name in src:
        if name not in ("xl/worksheets/sheet1.xml", "docProps/core.xml"):
            assert src[name] == out[name], name
    assert b'<c r="C5" t="inlineStr"><is><t xml:space="preserve">z</t>' in \
        out["xl/worksheets/sheet1.xml"]
    assert b"lastModifiedBy>Fixture Advisory LLP<" in out["docProps/core.xml"]
    assert_roundtrip(source, output, set(writes))


def test_the_proof_refuses_drift_and_a_changed_inventory(tmp_path):
    source = _unit_workbook(tmp_path / "src.xlsx")
    intended = {("S", "A1")}
    drifted = tmp_path / "drifted.xlsx"
    write_cells(source, drifted, {("S", "A1"): "x", ("S", "A2"): "moved"},
                firm=FIRM, at=AT)
    with pytest.raises(ContractError, match="drifted cell S!A2"):
        assert_roundtrip(source, drifted, intended)
    clean = tmp_path / "clean.xlsx"
    write_cells(source, clean, {("S", "A1"): "x"}, firm=FIRM, at=AT)
    assert_roundtrip(source, clean, intended)
    shrunk = tmp_path / "shrunk.xlsx"
    with zipfile.ZipFile(clean) as src, zipfile.ZipFile(shrunk, "w") as dst:
        for name in src.namelist():
            if name != "xl/styles.xml":
                dst.writestr(name, src.read(name))
    with pytest.raises(ContractError, match="changed the part inventory"):
        assert_roundtrip(source, shrunk, intended)
    with pytest.raises(ContractError, match="names sheet"):
        write_cells(source, tmp_path / "x.xlsx", {("Nope", "A1"): "x"},
                    firm=FIRM, at=AT)


def _pursuit(tmp_path) -> PursuitDir:
    """A hand-built client_provided pursuit over the writeback twin (the
    tests/web/test_writeback.py idiom): one drafted prose slot, one
    awaiting, on the questions sheet."""
    ws = tmp_path / "ws"
    pursuit = PursuitDir(ws, "pur_twin")
    (ws / "firm.json").write_text(json.dumps(FIRM), encoding="utf-8")
    source = pursuit.root / "inbox" / "writeback-twin.xlsx"
    source.write_bytes(TWIN.read_bytes())
    digest = hashlib.sha256(source.read_bytes()).hexdigest()

    def slot(slot_id, ref, cell):
        return {"slot_id": slot_id, "ref_id": ref,
                "source_mode": "client_provided", "response_shape": "prose",
                "fill_type": "authored",
                "source_locator": {"file": "writeback-twin.xlsx",
                                   "sheet": "1. Questions", "cell": cell}}

    container = {"pursuit_id": "pur_twin", "source_mode": "client_provided",
                 "parser_version": "test-1", "source_sha256": digest,
                 "slot_count": 2,
                 "slots": [slot("s_one", "1.0.1", "C2"),
                           slot("s_two", "1.0.2", "C3")]}
    pursuit.write_artifact("target_slots", container, name="slots.json")
    plant_freeze(pursuit, "pursuit_plan", {
        "pursuit_id": "pur_twin", "path": "A_designated",
        "slots_ref": "slots.json", "status": "approved",
        "sections": [{"section_id": "all", "slot_ids": ["s_one", "s_two"]}]})
    (pursuit.root / "drafts").mkdir(exist_ok=True)
    (pursuit.root / "drafts" / "draft.json").write_text(json.dumps({
        "plan_sha256": "0" * 64, "revision_n": 1,
        "sections": [{"section_id": "all", "answers": [
            {"slot_id": "s_one", "status": "drafted", "prose": PROSE},
            {"slot_id": "s_two", "status": "awaiting_disposition"}]}]}),
        encoding="utf-8")
    return pursuit


def test_the_twin_comes_back_byte_for_byte_plus_the_answer(tmp_path):
    """THE acceptance test: chart, drawing, comment, VML, validation and
    the cached formula value all survive; only the answered sheet part
    and core.xml differ, and the answer is in place."""
    pursuit = _pursuit(tmp_path)
    log = RunLogger(pursuit.root, pursuit.new_run_id(), pursuit.pursuit_id)
    log.run_start(mode="dry_run", engine_version=engine_version(),
                  config=effective_config(), kb_snapshot="kb@empty")
    facts = run_writeback(pursuit, log, at=AT, confirmed_by="Pat Lead")
    log.run_end(status="completed")
    validate("writeback_facts", facts)
    assert [c["decision"] for c in facts["cells"]] == ["written",
                                                       "empty_no_prose"]
    source = pursuit.root / facts["source_file"]
    output = pursuit.root / facts["output_file"]
    src, out = _members(source), _members(output)
    assert list(src) == list(out)
    for part in ("xl/charts/chart1.xml", "xl/drawings/drawing1.xml",
                 "xl/comments/comment1.xml", "xl/worksheets/sheet3.xml"):
        assert part in src, part
    changed = sorted(n for n in src if src[n] != out[n])
    assert changed == ["docProps/core.xml", "xl/worksheets/sheet2.xml"]
    assert any(n.endswith(".vml") for n in src), "the comment's VML part"
    # The cached value Excel wrote is still there for the next reader
    cached = load_workbook(output, data_only=True)
    assert cached["Pricing"]["B5"].value == 300
    assert load_workbook(output)["Pricing"]["B5"].value == "=SUM(B2:B4)"
    questions = load_workbook(output)["1. Questions"]
    assert questions["C2"].value == PROSE
    assert questions["C3"].value is None
    assert questions["C2"].fill.start_color.rgb == "FFFFFF00"
    assert questions["B2"].comment is not None
    assert b"lastModifiedBy>Fixture Advisory LLP<" in out["docProps/core.xml"]
    assert source.read_bytes() == TWIN.read_bytes(), "the inbox original"
