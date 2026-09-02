"""Write-back (B37/D19) — the frozen clause "write-back refuses pricing
cells": the template_fill pricing grid lands as refused_shape AND its
cells stay byte-identical (cell-model compare); the planted positive
proves a prose slot's cell IS written; formulas elsewhere survive; the
inbox original is never mutated; the facts record carries every
decision; an unconfirmed write refuses (operator required — S7)."""

import hashlib
import json

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from engine.web.server import create_app
from engine.workspace import PursuitDir
from tests.web.conftest import FIXED_AT, raising_caller, sign_in
from tests.helpers import plant_annotated, plant_freeze

ROLE = {"actor_role": "pursuit_lead"}


@pytest.fixture()
def wired(tmp_path):
    """A hand-built minimal pursuit: one prose slot (drafted), one
    template_fill pricing grid, one boolean, one unnamed-by-the-plan
    prose slot, one drafted-but-empty slot — plus a bystander formula.
    Every artifact goes through write_artifact (schema-validated)."""
    ws = tmp_path / "ws"
    pursuit = PursuitDir(ws, "pur_wb")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Response"
    sheet["A1"] = "Approach question"
    sheet["C1"] = ""            # the prose answer cell
    sheet["C2"] = "=SUM(D1:D9)"  # bystander formula (EC-5: must survive)
    sheet["C3"] = "$0"           # the pricing grid cell
    sheet["C4"] = "Yes/No"       # the boolean cell
    sheet["C5"] = ""             # unnamed-by-plan prose cell
    sheet["C6"] = ""             # drafted-but-empty (awaiting) cell
    source = pursuit.root / "inbox" / "buyer.xlsx"
    workbook.save(source)
    digest = hashlib.sha256(source.read_bytes()).hexdigest()

    def slot(slot_id, ref, shape, cell, **extra):
        return {"slot_id": slot_id, "ref_id": ref,
                "source_mode": "client_provided", "response_shape": shape,
                "source_locator": {"file": "buyer.xlsx",
                                   "sheet": "Response", "cell": cell},
                **extra}

    container = {
        "pursuit_id": "pur_wb", "source_mode": "client_provided",
        "parser_version": "test-1", "source_sha256": digest,
        "slot_count": 5,
        "slots": [
            slot("s_prose", "1.1", "prose", "C1", fill_type="authored"),
            slot("s_price", "9.1", "template_fill", "C3",
                 fill_type="template_fill"),
            slot("s_bool", "2.1", "boolean", "C4"),
            slot("s_unnamed", "3.1", "prose", "C5", fill_type="authored"),
            slot("s_empty", "4.1", "prose", "C6", fill_type="authored"),
        ],
    }
    pursuit.write_artifact("target_slots", container, name="slots.json")
    plan = {
        "pursuit_id": "pur_wb", "path": "A_designated",
        "slots_ref": "slots.json", "status": "approved",
        "sections": [{"section_id": "sec-1", "title": "Section 1",
                      # s_unnamed is deliberately NOT planned (T6)
                      "slot_ids": ["s_prose", "s_price", "s_bool",
                                   "s_empty"]}],
    }
    pursuit.write_artifact("pursuit_plan", plan, name="plan.json")
    frozen, _ = plant_freeze(pursuit, "pursuit_plan", plan, validate=True)
    envelope = {
        "pursuit_id": "pur_wb",
        "plan_sha256": hashlib.sha256(frozen.read_bytes()).hexdigest(),
        "revision_n": 0, "status": "complete",
        "sections": [{
            "section_id": "sec-1", "section_type": "methodology",
            "status": "drafted",
            "answers": [
                {"slot_id": "s_prose", "status": "drafted",
                 "prose": "Our phased cutover completes inside the "
                          "rehearsal-validated window."},
                {"slot_id": "s_empty", "status": "awaiting_disposition",
                 "reason": "gap undisposed"},
            ]}],
    }
    pursuit.write_artifact("draft", envelope, name="drafts/draft.json")
    plant_annotated(pursuit)
    app = create_app(ws, make_caller=raising_caller, now=lambda: FIXED_AT)
    with TestClient(app, base_url="http://127.0.0.1") as client:
        sign_in(client, "Wren Writer")
        yield client, pursuit, source


def test_writeback_refuses_pricing_cells(wired):
    client, pursuit, source = wired
    source_before = source.read_bytes()
    body = client.get("/api/pursuits/pur_wb/writeback/preview").json()
    preview = body["files"][0]  # the uniform shape (P18/C6, B77§2 D4)
    assert body["refused"] == []
    by_slot = {c["slot_id"]: c for c in preview["cells"]}
    assert by_slot["s_price"]["decision"] == "refused_shape"
    assert "template_fill" in by_slot["s_price"]["reason"]
    assert by_slot["s_bool"]["decision"] == "refused_shape"
    assert by_slot["s_unnamed"]["decision"] == "refused_unnamed"
    assert by_slot["s_empty"]["decision"] == "empty_no_prose"
    assert by_slot["s_prose"]["decision"] == "written"  # planted positive
    assert by_slot["s_prose"]["before"] == ""
    # the preview wrote NOTHING
    assert not (pursuit.root / "exports" / "writeback").exists()
    r = client.post("/api/pursuits/pur_wb/writeback/confirm",
                    json={"at": FIXED_AT, **ROLE})
    assert r.status_code == 200, r.text
    confirmed = r.json()
    facts = confirmed["files"][0]
    assert facts["confirmed_by"] == "Wren Writer"
    # the confirm composed the bundle: the filled form produced, the
    # rendered response recorded absent — never omitted
    by_lane = {d["lane"]: d for d in confirmed["bundle"]["deliverables"]}
    assert by_lane["xlsx_writeback"]["status"] == "produced"
    assert by_lane["submission_render"]["status"] == "absent"
    output = pursuit.root / "exports" / "writeback" / "buyer.xlsx"
    written = load_workbook(output)["Response"]
    # the prose landed; the refused cells are cell-model IDENTICAL
    assert written["C1"].value.startswith("Our phased cutover")
    assert written["C3"].value == "$0"        # pricing: untouched
    assert written["C4"].value == "Yes/No"    # boolean: untouched
    assert written["C5"].value is None        # unnamed: untouched
    assert written["C6"].value is None        # empty stays honestly empty
    # the bystander formula survived (EC-5)
    formulas = load_workbook(output)["Response"]
    assert formulas["C2"].value == "=SUM(D1:D9)"
    # the inbox ORIGINAL was never mutated (byte-identical)
    assert source.read_bytes() == source_before
    # the facts record landed schema-valid with the artifact lines
    stored = pursuit.read_artifact("exports/writeback-facts.json")
    assert stored["cells"] == facts["cells"]
    runs = sorted((pursuit.root / "runs").glob("*/run.jsonl"))
    records = json.loads("[" + ",".join(
        runs[-1].read_text().splitlines()) + "]")
    kinds = [r["artifact"]["kind"] for r in records
             if r.get("record_type") == "artifact"]
    assert kinds == ["writeback_facts", "write_back_file",
                     "submission_bundle"]


def test_unconfirmed_writeback_refuses(wired):
    client, pursuit, _ = wired
    bare = TestClient(client.app, base_url="http://127.0.0.1")  # no operator session: S7 unconfirmed
    assert bare.post("/api/pursuits/pur_wb/writeback/confirm",
                     json={"at": FIXED_AT, **ROLE}).status_code == 401
    assert not (pursuit.root / "exports" / "writeback").exists()


def test_writeback_refuses_wrong_source(wired, tmp_path):
    client, pursuit, source = wired
    source.write_bytes(b"tampered")  # the inbox file no longer matches
    r = client.get("/api/pursuits/pur_wb/writeback/preview")
    assert r.status_code == 409
    assert "source_sha256" in r.json()["detail"]