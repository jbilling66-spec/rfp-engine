"""The advisor (B37/D21): three-layer grounding (corpus-as-system with
the closed citation vocabulary, the discriminated reply union whose
decline arm carries no answer, the facts-only digest), the support cost
lane structurally unmixable from pursuit costs, and the decline
worklist. FakeCaller everywhere — conversational quality is a live-
milestone claim, never a CI one (B36(2))."""

import json

import pytest
from fastapi.testclient import TestClient

from engine.llm import FakeCaller
from engine.support import advisor as advisor_mod
from engine.support.advisor import (
    CITATION_VOCAB,
    DOC_SOURCES,
    compose_corpus,
    parse_reply,
    system_prompt,
)
from engine.web.server import create_app
from tests.validation.fixtures.validations import run_validation_package
from tests.web.conftest import FIXED_AT, raising_caller, sign_in


@pytest.fixture(scope="module")
def advised(tmp_path_factory):
    tmp = tmp_path_factory.mktemp("web-advisor")
    pursuit, report, _ = run_validation_package(tmp)
    assert report.status == "complete"
    app = create_app(tmp, make_caller=raising_caller, now=lambda: FIXED_AT)
    with TestClient(app, base_url="http://127.0.0.1") as client:
        sign_in(client, "Avery Asker")
        yield client, pursuit


def _script(reply: dict) -> FakeCaller:
    return FakeCaller({"advisor": json.dumps(reply)})


def test_corpus_is_deterministic_and_complete():
    corpus = compose_corpus()
    assert corpus == compose_corpus()
    for name in DOC_SOURCES:
        assert f"=== SOURCE: {name} ===" in corpus
    assert "(unavailable" not in corpus  # every doc really exists
    system = system_prompt()
    assert system.startswith("=== SOURCE:")
    assert system.rstrip().endswith("}`")  # the stable contract is LAST


def test_answer_flows_and_the_lane_is_unmixable(advised):
    client, pursuit = advised
    client.app.state.advisor_caller = _script(
        {"kind": "answer",
         "answer": "Press Advance on the pursuit detail screen.",
         "citations": ["pursuit-workflow.md", "not-a-real-source.md"]})
    runs_before = sorted((pursuit.root / "runs").glob("*"))
    r = client.post("/api/advisor", json={
        "question": "How do I run the next stage?",
        "pursuit_id": pursuit.pursuit_id})
    assert r.status_code == 200
    out = r.json()
    assert out["kind"] == "answer"
    assert out["citations"] == ["pursuit-workflow.md"]  # whitelist held
    # THE unmixable property: no run was created, no run-log line —
    # the spend lives in the support lane only, outside pursuits/
    assert sorted((pursuit.root / "runs").glob("*")) == runs_before
    workspace = pursuit.root.parent
    lines = [json.loads(l) for l in (
        workspace / "support" / "traces.jsonl").read_text().splitlines()]
    assert lines[-1]["outcome"] == "ok"
    assert lines[-1]["cost_usd"] > 0
    cost = client.get("/api/advisor/cost").json()
    assert cost["answered"] == 1 and cost["cost_source"] == "support_lane"


def test_decline_arm_carries_no_answer_and_feeds_the_worklist(advised):
    client, pursuit = advised
    client.app.state.advisor_caller = _script(
        {"kind": "not_covered", "topic": "invoice reconciliation",
         "answer": "smuggled!", "closest_sources": ["getting-started.md"]})
    r = client.post("/api/advisor", json={
        "question": "How do I reconcile the invoices?"})
    assert r.status_code == 200
    out = r.json()
    assert out["kind"] == "not_covered"
    assert "answer" not in out  # the union arm structurally cannot carry it
    r = client.post("/api/advisor", json={
        "question": "Invoices again?"})
    gaps = client.get("/api/advisor/gaps").json()
    assert gaps[0] == {"topic": "invoice reconciliation", "count": 2,
                       "last_at": FIXED_AT}


def test_advisor_guards(advised):
    client, pursuit = advised
    workspace = pursuit.root.parent
    # unknown pursuit: 404 WITHOUT creating a phantom directory
    r = client.post("/api/advisor", json={
        "question": "What about pur_ghost?", "pursuit_id": "pur_ghost"})
    assert r.status_code == 404
    assert not (workspace / "pur_ghost").exists()
    # a scalar wire (the P8 bug class) is a recorded ERROR, not a crash
    client.app.state.advisor_caller = FakeCaller({"advisor": "null"})
    r = client.post("/api/advisor", json={"question": "Anything?"})
    assert r.status_code == 502
    lines = [json.loads(l) for l in (
        workspace / "support" / "traces.jsonl").read_text().splitlines()]
    assert lines[-1]["outcome"] == "error"  # the failed call still counts
    # no session -> 401 (the advisor is firm-side, never guest-facing)
    bare = TestClient(client.app, base_url="http://127.0.0.1")
    assert bare.post("/api/advisor",
                     json={"question": "hello there"}).status_code == 401
    # the raw question text never lands in the lane — digest only
    assert "Anything?" not in (
        workspace / "support" / "traces.jsonl").read_text()


def test_missing_doc_becomes_a_marker_not_a_crash(monkeypatch, tmp_path):
    monkeypatch.setattr(advisor_mod, "CORPUS_DIR", tmp_path / "nowhere")
    corpus = advisor_mod.compose_corpus()
    assert corpus.count("(unavailable — this document is missing") == len(
        DOC_SOURCES)


def test_kb_export_is_read_only_and_clean(advised):
    client, pursuit = advised
    import io

    from openpyxl import load_workbook
    r = client.get("/api/kb/export.xlsx")
    assert r.status_code == 200
    sheet = load_workbook(io.BytesIO(r.content))["cards"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0][0] == "kb_id"
    assert len(rows) > 1  # the fixture store has cards
    # restricted provenance never exports: no original client strings
    payload = "\n".join(str(c) for row in rows for c in row)
    assert "source_client" not in payload
    assert "identifiers" not in payload
