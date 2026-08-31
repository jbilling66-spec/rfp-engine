"""Twin factory contract (P1): byte-deterministic builds matching the
committed goldens, with every baked-in edge case verifiable by reading the
workbook back."""

import hashlib
from pathlib import Path

import pytest
from openpyxl import load_workbook

from tests.fixtures.twins import GOLDENS, rebuild_all

FIXTURES_DIR = Path(__file__).resolve().parent


def _sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_builds_are_byte_deterministic(tmp_path):
    rebuild_all(tmp_path)
    first = {name: _sha(tmp_path / name) for name in GOLDENS}
    rebuild_all(tmp_path)
    assert {name: _sha(tmp_path / name) for name in GOLDENS} == first


@pytest.mark.parametrize("name", sorted(GOLDENS))
def test_rebuild_matches_committed_golden(tmp_path, name):
    golden = FIXTURES_DIR / name
    assert golden.exists(), f"golden {name} missing — run rebuild_all and commit it"
    GOLDENS[name](tmp_path / name)
    assert _sha(tmp_path / name) == _sha(golden), (
        f"{name} drifted from its committed golden — builder changes must be "
        "deliberate and the golden re-committed in the same change"
    )


def test_structured_twin_carries_its_edge_cases():
    wb = load_workbook(FIXTURES_DIR / "structured-twin.xlsx")
    assert "2. Integration " in wb.sheetnames  # EC-1 trailing space is real
    t = wb["2. Integration "]
    refs = [t[f"A{i}"].value for i in range(2, 6)]
    assert refs.count("2.0.5") == 2  # EC-2 duplicate ref preserved
    assert "Do not insert here" in t["C5"].value  # EC-6 appendix directive
    assert any(str(rng) == "D2:D4" for rng in t.merged_cells.ranges)  # EC-4
    assert t["B6"].value.startswith("='1. Company Background'!")  # EC-5a
    assert wb["3. Pricing"]["B5"].value == "=SUM(B2:B4)"  # EC-5b
    assert "250 words" in wb["Instructions"]["A2"].value  # EC-7 constraint
    assert wb["Instructions"]["A3"].value.startswith("*")  # EC-7 footnote


def test_nofill_twin_has_numbering_but_no_fills():
    """EC-3: the fixture that once produced zero slots silently. The P6
    parser acceptance runs against this file; here we prove the fixture
    actually has the trap property (questions present, no fill signal)."""
    wb = load_workbook(FIXTURES_DIR / "nofill-twin.xlsx")
    ws = wb["Scope of Services"]
    assert ws["A2"].value == "1.1" and ws["B5"].value  # questions exist
    fills = [
        ws.cell(row=r, column=c).fill.fill_type
        for r in range(1, 7)
        for c in range(1, 4)
    ]
    assert all(f in (None, "none") for f in fills)  # and nothing votes


def test_gapcase_twin_carries_its_trap_property():
    """P6 honest-gap fixture: sheet 1 asks answerable questions, sheet 2's
    questions carry the deliberately-off-corpus vocabulary. The disjointness
    against the LIVE corpus catalog is asserted in tests/planning (where the
    store exists); here we prove the planted questions are really present
    and the answer cells really vote (fill-mode parse)."""
    wb = load_workbook(FIXTURES_DIR / "gapcase-twin.xlsx")
    a = wb["1. Delivery Approach"]
    assert "data migration" in a["B2"].value
    assert "payroll parallel" in a["B3"].value
    b = wb["2. Special Requirements"]
    assert "quantum blockchain telemetry" in b["B2"].value
    assert "quantum blockchain telemetry" in b["B3"].value
    for ws in (a, b):
        for i in (2, 3):
            assert ws[f"C{i}"].fill.fill_type == "solid"  # answer cells vote


def test_twin_text_is_tripwire_clean():
    """Binary goldens dodge the text tripwire (suffix scan) — extract their
    text and sweep the real-client token list directly. New goldens ride
    this parametrization automatically via GOLDENS."""
    from engine.intake.extract import extract
    from tests.tripwire.tokens import scan_tokens

    tokens = scan_tokens()
    for name in sorted(GOLDENS):
        text = extract(FIXTURES_DIR / name).text.lower()
        for token in tokens:
            assert token not in text, f"{name} carries real-world token {token!r}"


def test_demo_twin_carries_every_validation_lane_carrier():
    """P8 demo package (B34(25)): every feature the validation stack needs
    non-vacuous must actually parse out of this file — the carriers are
    asserted here at the fixture, so a drifted twin fails before any
    downstream golden can silently stop exercising its lane."""
    wb = load_workbook(FIXTURES_DIR / "demo-twin.xlsx")
    ins = wb["Instructions"]
    assert "200 words" in ins["A2"].value                      # max_words carrier
    assert "state plainly that it is not offered" in ins["A3"].value
    assert "partners deliver" in ins["A4"].value               # disclose flag
    assert "appendix" not in " ".join(
        str(ins[f"A{i}"].value) for i in range(1, 5)).lower()  # no global appendix stamp

    assert wb["7. Subcontracting"]["B2"].value.startswith("Do ")   # boolean opener
    assert "If yes to 7.0.1" in wb["7. Subcontracting"]["B3"].value  # gated child
    assert wb["2. Data Migration"]["B4"].value.count("4.0.1") == 1   # cross_ref
    uat = wb["4. Testing"]["B3"].value
    assert uat.count("?") >= 2                                  # sub_questions
    assert "Do not insert here" in wb["8. Appendices"]["C2"].value  # appendix EC-6
    assert "separate appendix" in wb["8. Appendices"]["B3"].value
    assert "decommissioning on-premises" in wb["8. Appendices"]["B4"].value  # off-corpus gap carrier

    question_sheets = [n for n in wb.sheetnames if n != "Instructions"]
    assert len(question_sheets) == 8
    filled = sum(
        1
        for name in question_sheets
        for row in wb[name].iter_rows(min_row=2, min_col=3, max_col=3)
        for cell in row
        if cell.fill.fill_type == "solid"
    )
    assert filled == 19                                         # every slot votes
