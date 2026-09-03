"""Byte-deterministic synthetic ERP twin workbooks (P1).

Fresh implementations for a synthetic buyer ("Northwind Regional Health",
ERP implementation services) that BAKE IN the edge cases v1 paid for in
production, each tagged [EC-n] and referenced from docs/v1-ledger.md:

  EC-1  trailing-space sheet name — byte-exact names or write-back misses
  EC-2  duplicate ref_id (2.0.5 twice) — preserved, never "fixed"
  EC-3  a no-fill workbook numbered 1.1-style — once parsed to ZERO slots
        silently; the worst failure for a zero-silent-misses engine
  EC-4  merged criterion cells — values forward-fill from merged ranges
  EC-5  cross-sheet formula in an ordinary title cell + a live =SUM() in a
        grid — must survive write-back untouched
  EC-6  appendix-directive answer cell ("Do not insert here…") — an
        appendix-routed LEAF, not a record
  EC-7  length constraint living in an Instructions sheet + a footnote
        that is a sheet-scoped global constraint
  EC-8  openpyxl re-stamps dcterms:modified from the wall clock inside
        save() — byte determinism requires rebuilding the archive with
        pinned timestamps and stable member order

Determinism contract: build_* writes the same bytes every run; the goldens
committed beside this file are compared byte-for-byte in tests.
"""

import io
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

ANSWER_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
_PINNED_DATE = (2026, 1, 1, 0, 0, 0)  # EC-8: fixed zip entry timestamp


def _freeze(workbook: Workbook, path: Path) -> None:
    """EC-8: save via openpyxl, then rebuild the zip with pinned entry
    timestamps, stable member order, and pinned core properties."""
    import datetime as _dt
    pinned = _dt.datetime(2026, 1, 1)
    workbook.properties.created = pinned
    workbook.properties.modified = pinned
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    with zipfile.ZipFile(buffer) as src, zipfile.ZipFile(
        path, "w", zipfile.ZIP_DEFLATED
    ) as dst:
        for name in sorted(src.namelist()):
            data = src.read(name)
            if name == "docProps/core.xml":
                # EC-8 proper: openpyxl re-stamps dcterms:modified from the
                # wall clock INSIDE save(), ignoring the pinned property.
                # Replace only the element TEXT — the namespace declarations
                # live on the element's own attributes and must survive.
                import re
                data = re.sub(
                    rb"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)",
                    rb"\g<1>2026-01-01T00:00:00Z\g<2>",
                    data,
                )
            info = zipfile.ZipInfo(name, date_time=_PINNED_DATE)
            info.compress_type = zipfile.ZIP_DEFLATED
            dst.writestr(info, data)


def build_structured_twin(path: Path) -> Path:
    """The Path-A questionnaire twin: an ERP implementation services
    workbook with designated answer cells."""
    _freeze(_structured_workbook(), path)
    return path


def _structured_workbook() -> Workbook:
    """The structured twin's workbook, unfrozen — shared with the formula
    twin (P1-24) so the two differ ONLY by the planted cached value."""
    wb = Workbook()

    # Instructions sheet with a length constraint + sheet-scoped footnote (EC-7)
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "Northwind Regional Health — ERP Implementation Services RFP"
    ws["A2"] = "Respond in the yellow cells only. Limit narrative responses to 250 words."
    ws["A3"] = "* Rate-card figures on the Pricing sheet must reflect blended onshore rates."

    # Sheet 1 — numbered questions, ref column + answer column
    q = wb.create_sheet("1. Company Background")
    q["A1"] = "Ref"
    q["B1"] = "Question"
    q["C1"] = "Response"
    rows = [
        ("1.0.1", "Provide an overview of your firm and ERP practice."),
        ("1.0.2", "Describe your implementation methodology."),
        ("1.0.3", "Summarize comparable implementations for regional health systems."),
    ]
    for i, (ref, question) in enumerate(rows, start=2):
        q[f"A{i}"] = ref
        q[f"B{i}"] = question
        q[f"C{i}"] = ""
        q[f"C{i}"].fill = ANSWER_FILL

    # Sheet 2 — trailing space in the NAME (EC-1); duplicate ref_id (EC-2);
    # appendix directive (EC-6); merged criterion cells (EC-4)
    t = wb.create_sheet("2. Integration ")  # trailing space is REAL
    t["A1"] = "Ref"
    t["B1"] = "Question"
    t["C1"] = "Response"
    t["D1"] = "Criterion"
    data = [
        ("2.0.1", "Describe your integration approach for HL7/FHIR interfaces."),
        ("2.0.5", "List integration accelerators for payroll systems."),
        ("2.0.5", "List integration accelerators for scheduling systems."),  # EC-2
        ("2.0.7", "Provide your integration governance RACI."),
    ]
    for i, (ref, question) in enumerate(data, start=2):
        t[f"A{i}"] = ref
        t[f"B{i}"] = question
    for i in (2, 3, 4):
        t[f"C{i}"] = ""
        t[f"C{i}"].fill = ANSWER_FILL
    t["C5"] = "Do not insert here. Include as part of the Integration RACI Appendix."
    t["C5"].fill = ANSWER_FILL  # EC-6: directive text living in the answer cell
    t.merge_cells("D2:D4")  # EC-4
    t["D2"] = "Technical capability (30%)"

    # EC-5a: cross-sheet formula in an ordinary title cell
    t["B6"] = "='1. Company Background'!B2"

    # Sheet 3 — pricing grid with a live formula (EC-5b); pricing is
    # human-only in v2, so write-back must refuse this sheet entirely
    p = wb.create_sheet("3. Pricing")
    p["A1"] = "Phase"
    p["B1"] = "Hours"
    for i, (phase, hours) in enumerate(
        [("Discover", 400), ("Configure", 900), ("Deploy", 600)], start=2
    ):
        p[f"A{i}"] = phase
        p[f"B{i}"] = hours
    p["B5"] = "=SUM(B2:B4)"  # EC-5b: must survive write-back untouched
    return wb


B6_CACHED = "Provide an overview of your firm and ERP practice."  # = sheet 1 B2
_B6_EMPTY = b"<c r=\"B6\"><f>'1. Company Background'!B2</f><v></v></c>"
_B6_CACHED = (b"<c r=\"B6\" t=\"str\"><f>'1. Company Background'!B2</f><v>"
              + B6_CACHED.encode() + b"</v></c>")


def build_formula_twin(path: Path) -> Path:
    """P1-24 (P26b-1, B112): the structured twin with row 6 of sheet 2
    made a QUESTION ROW (ref 2.0.9) whose question is the cross-sheet
    formula — and, spliced into the sheet XML after the save, the cached
    value Excel would have stored (openpyxl cannot write one). The
    parser must emit the slot from the cache; the structured twin, which
    has no cache, must warn instead."""
    wb = _structured_workbook()
    wb["2. Integration "]["A6"] = "2.0.9"
    _freeze(wb, path)
    _splice_member(path, "xl/worksheets/sheet3.xml", _B6_EMPTY, _B6_CACHED)
    return path


def _splice_member(path: Path, member: str, old: bytes, new: bytes) -> None:
    """Rewrite one zip member's bytes, keeping EC-8's pinned timestamps
    and member order (the frozen container, edited in place)."""
    buffer = io.BytesIO(path.read_bytes())
    with zipfile.ZipFile(buffer) as src, zipfile.ZipFile(
        path, "w", zipfile.ZIP_DEFLATED
    ) as dst:
        for name in sorted(src.namelist()):
            data = src.read(name)
            if name == member:
                assert data.count(old) == 1, (member, data.count(old))
                data = data.replace(old, new)
            info = zipfile.ZipInfo(name, date_time=_PINNED_DATE)
            info.compress_type = zipfile.ZIP_DEFLATED
            dst.writestr(info, data)


def build_nofill_twin(path: Path) -> Path:
    """EC-3: a services workbook with numbering but NO answer fills anywhere.
    v1's conventions-learner had nothing to vote with, defaulted its leaf
    depth, and every question fell between two rules — zero slots, silently.
    The v2 parser must either find these slots structurally or refuse loudly;
    an empty parse of this file is the named regression."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Scope of Services"
    ws["A1"] = "Section"
    ws["B1"] = "Requirement"
    rows = [
        ("1.1", "Vendor shall describe its project management approach."),
        ("1.2", "Vendor shall provide a staffing plan with named roles."),
        ("2.1", "Vendor shall describe its testing methodology."),
        ("2.2", "Vendor shall describe its training and OCM approach."),
    ]
    for i, (ref, requirement) in enumerate(rows, start=2):
        ws[f"A{i}"] = ref
        ws[f"B{i}"] = requirement
    _freeze(wb, path)
    return path


def build_gapcase_twin(path: Path) -> Path:
    """The honest-gap twin (P6): a designated questionnaire where sheet 1's
    questions are answerable from the seeded tests/kb corpus and sheet 2's
    are deliberately OFF-corpus — every content token of the sheet-2
    questions is absent from every card's catalog text (title/summary/tags),
    so card_search returns [] and the KB Mapper must flag honest no_content
    gaps instead of inventing. The vocabulary-disjointness property is
    asserted against the live corpus in tests/planning (non-vacuous by
    construction: if a future corpus card adopts these tokens, that test
    fails before this twin can silently stop gapping)."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "Northwind Regional Health — ERP Implementation Services Questionnaire (Addendum)"
    ws["A2"] = "Respond in the yellow cells only."

    a = wb.create_sheet("1. Delivery Approach")
    a["A1"] = "Ref"
    a["B1"] = "Question"
    a["C1"] = "Response"
    on_corpus = [
        ("1.0.1", "Describe the data migration approach for patient and financial records."),
        ("1.0.2", "Describe the testing methodology, including payroll parallel runs."),
    ]
    for i, (ref, question) in enumerate(on_corpus, start=2):
        a[f"A{i}"] = ref
        a[f"B{i}"] = question
        a[f"C{i}"] = ""
        a[f"C{i}"].fill = ANSWER_FILL

    b = wb.create_sheet("2. Special Requirements")
    b["A1"] = "Ref"
    b["B1"] = "Question"
    b["C1"] = "Response"
    # Prose-shaped on purpose: a boolean/numeric opener would shape-skip
    # the slot instead of gapping it, defeating the trap.
    off_corpus = [
        ("2.0.1", "Provide quantum blockchain telemetry certification evidence for proposed personnel."),
        ("2.0.2", "Outline anticipated quantum blockchain telemetry uptime guarantees."),
    ]
    for i, (ref, question) in enumerate(off_corpus, start=2):
        b[f"A{i}"] = ref
        b[f"B{i}"] = question
        b[f"C{i}"] = ""
        b[f"C{i}"].fill = ANSWER_FILL

    _freeze(wb, path)
    return path


def build_demo_twin(path: Path) -> Path:
    """The M1 slice demo package (P8, B34(25)): the workbook rich enough for
    a live drafting-quality review — 8 question sheets spanning six
    corpus-saturated section_type families, ~18 slots, and one carrier for
    every validation lane P8 must exercise non-vacuously: a real word limit
    + state_if_not_offered + disclose_partner_delivery in the Instructions
    (global constraints stamp every slot), a boolean gate with a gated
    child, a sub_questions slot, a cross_refs pair, and appendix-routed
    slots. Same synthetic buyer as every twin (tripwire-safe)."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "Northwind Regional Health — ERP Implementation Services RFP (M1 Demonstration Package)"
    ws["A2"] = "Respond in the yellow cells only. Limit narrative responses to 200 words."
    ws["A3"] = "If a service or capability is not offered, state plainly that it is not offered."
    ws["A4"] = "Disclose in your response any services your delivery partners deliver on your behalf."

    def sheet(title, rows, *, directive_rows=()):
        s = wb.create_sheet(title)
        s["A1"], s["B1"], s["C1"] = "Ref", "Question", "Response"
        for i, (ref, question) in enumerate(rows, start=2):
            s[f"A{i}"] = ref
            s[f"B{i}"] = question
            s[f"C{i}"] = directive_rows[i - 2] if i - 2 < len(directive_rows) and directive_rows[i - 2] else ""
            s[f"C{i}"].fill = ANSWER_FILL
        return s

    sheet("1. Implementation Methodology", [
        ("1.0.1", "Describe your implementation methodology for ERP deployments."),
        ("1.0.2", "Describe your project governance approach, including steering cadence."),
        ("1.0.3", "Summarize your delivery timeline approach and key milestones."),
    ])
    sheet("2. Data Migration", [
        ("2.0.1", "Describe the data migration approach for patient and financial records."),
        ("2.0.2", "Describe your reconciliation and trial load approach for legacy data."),
        ("2.0.3", "Describe your cutover rehearsal approach and confirm it remains "
                  "consistent with the testing methodology described in 4.0.1."),
    ])
    sheet("3. Integration", [
        ("3.0.1", "Describe your integration approach for HL7/FHIR interfaces."),
        ("3.0.2", "List integration accelerators for payroll systems."),
    ])
    sheet("4. Testing", [
        ("4.0.1", "Describe the testing methodology, including payroll parallel runs."),
        ("4.0.2", "Describe your user acceptance testing approach. How are defects "
                  "triaged during test cycles? How do you determine exit criteria "
                  "for each phase?"),
    ])
    sheet("5. Training and OCM", [
        ("5.0.1", "Describe your end-user training approach for clinical and back-office staff."),
        ("5.0.2", "Describe your organizational change management approach."),
    ])
    sheet("6. Support Model", [
        ("6.0.1", "Describe your post-go-live support model, including hypercare."),
        ("6.0.2", "Describe your issue escalation and severity management approach."),
    ])
    sheet("7. Subcontracting", [
        ("7.0.1", "Do you use delivery partners or subcontractors for implementation services?"),
        ("7.0.2", "If yes to 7.0.1, describe the services your partners deliver and "
                  "where those teams are located."),
    ])
    sheet("8. Appendices", [
        ("8.0.1", "Provide your integration governance RACI."),
        ("8.0.2", "Provide implementation reference architecture diagrams as a separate appendix."),
        ("8.0.3", "Describe your capabilities for decommissioning on-premises "
                  "mainframes, including COBOL source translation."),
    ], directive_rows=(
        "Do not insert here. Include as part of the Integration RACI Appendix.",
        "",
        "",
    ))

    _freeze(wb, path)
    return path


GOLDENS = {
    "structured-twin.xlsx": build_structured_twin,
    "nofill-twin.xlsx": build_nofill_twin,
    "gapcase-twin.xlsx": build_gapcase_twin,
    "demo-twin.xlsx": build_demo_twin,
    "formula-twin.xlsx": build_formula_twin,  # P1-24 (P26b-1)
}


def rebuild_all(directory: Path) -> None:
    for name, builder in GOLDENS.items():
        builder(directory / name)
