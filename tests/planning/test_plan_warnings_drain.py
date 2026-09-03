"""P1-23 (P26b-1, B112): the planner drains parser warnings into the plan
report AND the run log — one recoverable `error` record per file that
carried any, none for a clean parse. The structured twin's single
warning is hand-derived in tests/structure/test_parse_warnings.py.
"""

from openpyxl import Workbook

from engine.runlog import read_run
from tests.planning.fixtures.plans import run_planning_package


def _parser_records(pursuit):
    records = read_run(pursuit.root / "runs" / pursuit.latest_run_id()
                       / "run.jsonl")
    return [r for r in records if r["record_type"] == "error"
            and r.get("error", {}).get("code") == "parser_warnings"]


def test_structured_twin_warning_reaches_report_and_run_log(tmp_path):
    pursuit, report = run_planning_package(tmp_path, package_id="xlsx",
                                           gate2=None)
    assert report.status == "complete"
    line = ("structured-twin.xlsx: 2. Integration!row 6: formula-only row "
            "skipped (a formula question needs a cached value)")
    assert line in report.warnings
    recs = _parser_records(pursuit)
    assert len(recs) == 1
    err = recs[0]["error"]
    assert err["recoverable"] is True
    assert err["action_taken"] == "surfaced_to_human"
    # Two lines ride ONE record: sheet 2 row 6 (the cross-sheet formula)
    # and sheet 3 row 5 (the =SUM total in the pricing grid).
    assert err["message"].startswith("structured-twin.xlsx: 2 parser warning(s)")
    assert "2. Integration!row 6" in err["message"]
    assert "3. Pricing!row 5" in err["message"]


def test_a_clean_parse_drains_nothing(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws["A1"], ws["B1"], ws["C1"] = "Ref", "Question", "Response"
    ws["A2"], ws["B2"] = "1.1", "Describe your approach."
    ws["A3"], ws["B3"] = "1.2", "Describe your team."
    ws["A4"], ws["B4"] = "1.3", "Describe your pricing model."
    path = tmp_path / "clean-twin.xlsx"
    wb.save(path)
    pursuit, report = run_planning_package(tmp_path, package_id="xlsx",
                                           workbook=path, gate2=None)
    assert report.status == "complete"
    assert not [w for w in report.warnings if "clean-twin.xlsx" in w]
    assert _parser_records(pursuit) == []
