"""Intake fixture goldens: byte-deterministic, honestly planted, tripwire-clean.

Raw-library assertions only (pypdf/openpyxl) — engine extraction is covered
by test_extract.py. This file proves the FIXTURES: the committed bytes match
a rebuild, the planted content is really there (and the injection sentence
ONLY in the injected variant), and — because the tripwire scans text files,
not binaries — the extractable text of every binary golden is clean of the
tracked client tokens.
"""

from pathlib import Path

import pypdf
import pytest
from openpyxl import load_workbook

from tests.fixtures.intake_twins import (
    GOLDENS,
    HIDDEN_ROW_DIRECTIVE,
    HIDDEN_SHEET_DIRECTIVE,
    INJECTION_SENTENCE,
)
from tests.tripwire.tokens import scan_tokens

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"


def _pdf_text(path: Path) -> str:
    return "\n".join(page.extract_text() for page in pypdf.PdfReader(path).pages)


def _xlsx_text(path: Path) -> str:
    wb = load_workbook(path)
    cells = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells.extend(str(v) for v in row if v is not None)
    return "\n".join(cells)


@pytest.mark.parametrize("name", sorted(GOLDENS))
def test_rebuild_matches_committed_golden(name, tmp_path):
    rebuilt = GOLDENS[name](tmp_path / name)
    assert rebuilt.read_bytes() == (FIXTURES / name).read_bytes(), (
        f"{name} drifted from its builder — rerun intake_twins.rebuild_all "
        "and commit the binary with the change that moved it"
    )


@pytest.mark.parametrize("name", sorted(GOLDENS))
def test_build_is_deterministic(name, tmp_path):
    a = GOLDENS[name](tmp_path / f"a-{name}").read_bytes()
    b = GOLDENS[name](tmp_path / f"b-{name}").read_bytes()
    assert a == b


def test_pdf_twin_planted_content_present():
    reader = pypdf.PdfReader(FIXTURES / "pdf-twin.pdf")
    page1 = reader.pages[0].extract_text()
    page2 = reader.pages[1].extract_text()
    # deadlines + submission (page 1)
    assert "August 8, 2026" in page1
    assert "August 29, 2026 at 3:00 PM Central Time" in page1
    assert "Northwind procurement portal" in page1
    assert "Form A (Vendor Certification)" in page1
    # weights table + shall-requirements (page 2)
    for line in ["Technical approach: 40%", "Implementation team: 25%",
                 "Past performance: 20%", "Price: 15%"]:
        assert line in page2
    assert "3.1 Vendor shall" in page2
    # AI-disclosure and OCI clauses (page 2)
    assert "generative AI" in page2
    assert "original work" in page2
    assert "organizational conflict of interest" in page2
    assert "audit or attestation relationship" in page2


def test_injection_sentence_only_in_injected_variant():
    clean = _pdf_text(FIXTURES / "pdf-twin.pdf")
    injected = _pdf_text(FIXTURES / "pdf-twin-injected.pdf")
    assert INJECTION_SENTENCE not in clean
    assert INJECTION_SENTENCE in injected
    # the variants differ by exactly the planted sentence
    assert injected.replace(INJECTION_SENTENCE, "").replace("\n", "") == clean.replace("\n", "")


def test_hidden_twin_content_is_actually_hidden():
    wb = load_workbook(FIXTURES / "hidden-twin.xlsx")
    assert wb["Internal Notes"].sheet_state == "hidden"
    assert wb["Internal Notes"]["A1"].value == HIDDEN_SHEET_DIRECTIVE
    ws = wb["Vendor Questions"]
    assert ws.row_dimensions[4].hidden is True
    assert ws["B4"].value == HIDDEN_ROW_DIRECTIVE


@pytest.mark.parametrize("name", sorted(GOLDENS))
def test_binary_fixture_text_is_tripwire_clean(name):
    path = FIXTURES / name
    text = (_pdf_text(path) if name.endswith(".pdf") else _xlsx_text(path)).lower()
    for token in scan_tokens():
        assert token not in text, f"tripwire token {token!r} inside binary fixture {name}"
