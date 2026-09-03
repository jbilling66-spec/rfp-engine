"""P1-25 (P26b-1, B112): intake marks every formula cell, renders the
cached value beside the source when the writer saved one, and warns —
by address, never by text — when it did not. The source text itself is
kept (EC-5: never thinned). The formula twin's cache is hand-derived in
`tests/fixtures/twins.py` (B6_CACHED)."""

from pathlib import Path

from openpyxl import Workbook

from engine.intake.extract import extract
from tests.fixtures.twins import B6_CACHED

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"


def test_cached_value_renders_beside_the_marked_source():
    doc = extract(FIXTURES / "formula-twin.xlsx")
    assert f"[formula ='1. Company Background'!B2 → {B6_CACHED}]" in doc.text
    assert not any("2. Integration !B6" in w for w in doc.warnings)


def test_missing_cache_is_marked_and_warned_by_address():
    doc = extract(FIXTURES / "structured-twin.xlsx")
    assert "[formula, no cached value] ='1. Company Background'!B2" in doc.text
    assert "[formula, no cached value] =SUM(B2:B4)" in doc.text
    addresses = [w.split(":")[0] for w in doc.warnings if "formula cell" in w]
    assert addresses == ["2. Integration !B6", "3. Pricing!B5"]
    assert all("SUM" not in w and "Company" not in w for w in doc.warnings)


def test_a_workbook_without_formulas_warns_nothing(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = "Ref", "Question"
    ws["A2"], ws["B2"] = "1.1", "Describe your approach."
    path = tmp_path / "plain.xlsx"
    wb.save(path)
    doc = extract(path)
    assert "[formula" not in doc.text
    assert not [w for w in doc.warnings if "formula" in w]
