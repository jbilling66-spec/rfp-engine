"""P1-26 (P26b-1, B112): hidden COLUMNS are extracted, marked in the
text, and recorded as their own hidden segment — one per column, located
by sheet and letter — so the hidden_content finding and the injection
screen see them exactly as they see hidden sheets and rows."""

from pathlib import Path

from openpyxl import Workbook

from engine.intake.extract import extract
from tests.fixtures.intake_twins import HIDDEN_COLUMN_DIRECTIVE

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"


def test_hidden_twin_column_is_marked_and_segmented():
    doc = extract(FIXTURES / "hidden-twin.xlsx")
    assert f"[hidden col] {HIDDEN_COLUMN_DIRECTIVE}" in doc.text
    seg = next(s for s in doc.hidden_segments
               if s["location"] == "hidden-twin.xlsx: Vendor Questions!C")
    assert seg["text"] == HIDDEN_COLUMN_DIRECTIVE
    # Hand count: the hidden row (Vendor Questions), the hidden sheet's
    # one row (Internal Notes), the hidden column — three segments.
    assert len(doc.hidden_segments) == 3


def test_a_hidden_range_of_columns_is_expanded(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"], ws["B1"], ws["C1"], ws["D1"] = "keep", "one", "two", "keep"
    ws["B2"], ws["C2"] = "three", "four"
    dim = ws.column_dimensions["B"]
    dim.min, dim.max, dim.hidden = 2, 3, True  # B:C hidden as one range
    path = tmp_path / "range.xlsx"
    wb.save(path)
    doc = extract(path)
    by_loc = {s["location"]: s["text"] for s in doc.hidden_segments}
    assert by_loc == {"range.xlsx: S!B": "one three", "range.xlsx: S!C": "two four"}
    assert "[hidden col] keep" not in doc.text
