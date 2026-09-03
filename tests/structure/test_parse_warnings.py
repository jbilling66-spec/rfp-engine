"""P1-23 (P26b-1, B112): the parser's warnings carrier.

Every row the workbook parser dropped while it carried content is a
recorded warning on `ParsedWorkbook.warnings` — sheet, row and kind,
never a cell's text. Expectations are hand-derived from the fixture
builders (`tests/fixtures/twins.py`), never pasted from a run.
"""

import re
from pathlib import Path

from openpyxl import Workbook

from engine.structure import parse_workbook
from engine.structure.parse import ParsedWorkbook

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"
TWINS = ("structured-twin.xlsx", "nofill-twin.xlsx", "gapcase-twin.xlsx",
         "demo-twin.xlsx")
_LINE = re.compile(r"^.+!row \d+: [a-z-]+[^\n]*$")


def test_every_twin_carries_the_field_and_every_line_is_shaped():
    for name in TWINS:
        parsed = parse_workbook(FIXTURES / name)
        assert isinstance(parsed, ParsedWorkbook)
        assert isinstance(parsed.warnings, list), name
        for line in parsed.warnings:
            assert _LINE.match(line), (name, line)


def test_structured_twin_records_its_formula_only_row():
    # Hand-derived: sheet "2. Integration " row 6 holds ONLY the EC-5a
    # cross-sheet formula in B6 (no ref, no answer cell), so it matches
    # no rule and is the one row the terminal drop sees. Sheet 3's TOTAL
    # row (B5 = SUM, nothing else on the row) leaves through the grid
    # branch. The Instructions sheet is never classified.
    parsed = parse_workbook(FIXTURES / "structured-twin.xlsx")
    assert parsed.warnings == [
        "2. Integration!row 6: formula-only row skipped (a formula "
        "question needs a cached value)",
        "3. Pricing!row 5: formula-only grid row skipped (totals are "
        "facts, never targets)",
    ]
    assert parsed.slot_count == 8  # the warning is additive — nothing moved


def test_warnings_never_carry_cell_text(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws["A1"], ws["B1"], ws["C1"] = "Ref", "Question", "Response"
    ws["A2"], ws["B2"] = "1.1", "Describe your approach."
    ws["A3"], ws["B3"] = "1.2", "Describe your team."
    # Text in the ANSWER column of a row with no ref and no question:
    # no rule matches it. (A no-ref row with text in the question
    # column is a legitimate un-numbered question and parses as a
    # leaf; a row with two empty labeled columns parses as a record
    # ask — neither is this case.)
    ws["C5"] = "CONFIDENTIAL BANNER TEXT — do not distribute"
    path = tmp_path / "banner.xlsx"
    wb.save(path)
    parsed = parse_workbook(path)
    assert any(line.startswith("Questions!row 5: ") for line in parsed.warnings)
    assert all("CONFIDENTIAL" not in line for line in parsed.warnings)


def test_a_clean_questionnaire_has_no_warnings(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws["A1"], ws["B1"], ws["C1"] = "Ref", "Question", "Response"
    ws["A2"], ws["B2"] = "1.1", "Describe your approach."
    ws["A3"], ws["B3"] = "1.2", "Describe your team."
    path = tmp_path / "clean.xlsx"
    wb.save(path)
    assert parse_workbook(path).warnings == []


def test_a_ref_row_above_the_header_is_recorded(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws["A1"], ws["B1"] = "0.1", "Confirm receipt of this addendum."  # above the header
    ws["A2"], ws["B2"], ws["C2"] = "Ref", "Question", "Response"
    ws["A3"], ws["B3"] = "1.1", "Describe your approach."
    ws["A4"], ws["B4"] = "1.2", "Describe your team."
    path = tmp_path / "preamble.xlsx"
    wb.save(path)
    parsed = parse_workbook(path)
    assert ("Questions!row 1: row above the first label row skipped "
            "(carries a ref)") in parsed.warnings
    assert all("addendum" not in line for line in parsed.warnings)
