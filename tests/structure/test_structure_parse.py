"""Path-A parser acceptance: the three twins parse to their hand-
transcribed goldens, and every produced slot validates against the
frozen target-slot bench contract (its first instance-level coverage).

Goldens are transcribed by reading the committed twins directly (house
pattern), never by running the parser and pasting its output.
"""

from pathlib import Path

from openpyxl import Workbook

from engine.contracts import validate
from engine.structure import PARSER_VERSION, parse_workbook

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"


def test_every_slot_validates_against_frozen_schema():
    for name in ("structured-twin.xlsx", "nofill-twin.xlsx", "gapcase-twin.xlsx"):
        parsed = parse_workbook(FIXTURES / name)
        assert parsed.slots, name
        for slot in parsed.slots:
            validate("target_slot", slot)


def test_parse_is_deterministic():
    a = parse_workbook(FIXTURES / "structured-twin.xlsx")
    b = parse_workbook(FIXTURES / "structured-twin.xlsx")
    assert a.slots == b.slots
    assert a.source_sha256 == b.source_sha256
    assert a.parser_version == PARSER_VERSION


def test_structured_twin_matches_golden():
    parsed = parse_workbook(FIXTURES / "structured-twin.xlsx")
    by_id = {s["slot_id"]: s for s in parsed.slots}

    # Hand count: 3 leaves on sheet 1, 4 on sheet 2, 1 grid slot. No headers.
    assert parsed.slot_count == 8
    assert [s.get("ref_id") for s in parsed.slots] == [
        "1.0.1", "1.0.2", "1.0.3", "2.0.1", "2.0.5", "2.0.5", "2.0.7", None,
    ]  # EC-2: the duplicate 2.0.5 is preserved, never "fixed"

    # EC-1: sheet name keeps its trailing space in the locator, stripped
    # in the human path.
    s = by_id["slot_02_r002"]
    assert s["source_locator"]["sheet"] == "2. Integration "
    assert s["path"] == "2. Integration > 2.0.1"

    # EC-4: merged criterion forward-fills rows 2-4; row 5 is outside.
    for sid in ("slot_02_r002", "slot_02_r003", "slot_02_r004"):
        assert by_id[sid]["eval_criterion"] == "Technical capability (30%)"
    assert "eval_criterion" not in by_id["slot_02_r005"]

    # EC-6: the appendix directive in the answer cell routes the leaf.
    appendix = by_id["slot_02_r005"]
    assert appendix["answer_location"] == "appendix"
    assert appendix["appendix_ref"] == "Integration Raci Appendix"
    assert appendix["question_text"].startswith("Provide your integration governance")

    # EC-5: neither formula cell (cross-sheet title B6, =SUM total B5)
    # produced a slot. Still true under P1-24 (2.1.0): this twin is
    # openpyxl-built, so B6 carries NO cached value — it is recorded as a
    # parser warning instead; formula-twin.xlsx is the slot-emitting case.
    assert not any(
        s["source_locator"].get("cell") in ("B6", "B5") for s in parsed.slots
    )

    # The pricing grid collapsed to ONE template_fill slot with typed
    # fields from the real header labels.
    grid = by_id["slot_03_r001"]
    assert grid["response_shape"] == "template_fill"
    assert grid["fill_type"] == "template_fill"
    assert [(f["key"], f["type"]) for f in grid["response_fields"]] == [
        ("phase", "text"), ("hours", "number"),
    ]


def test_nofill_twin_parses_four_slots_structurally():
    """EC-3: the named regression — numbering but no fills once parsed to
    ZERO slots silently. The structural fallback must find all four."""
    parsed = parse_workbook(FIXTURES / "nofill-twin.xlsx")
    assert [s["ref_id"] for s in parsed.slots] == ["1.1", "1.2", "2.1", "2.2"]
    for slot in parsed.slots:
        assert slot["response_shape"] == "prose"
        assert slot["source_locator"]["sheet"] == "Scope of Services"


def test_gapcase_twin_parses_all_four_as_prose():
    """The honest-gap twin's off-corpus asks must stay prose — a boolean
    or numeric shape would be shape-skipped by the mapper instead of
    gapped, silently defeating the trap."""
    parsed = parse_workbook(FIXTURES / "gapcase-twin.xlsx")
    assert [s["ref_id"] for s in parsed.slots] == ["1.0.1", "1.0.2", "2.0.1", "2.0.2"]
    assert all(s["response_shape"] == "prose" for s in parsed.slots)


def test_header_slots_and_gating_link(tmp_path):
    """Rule 3 (ref shallower than leaf depth -> header slot, no response)
    and the explicit gating link, on a purpose-built workbook."""
    from tests.fixtures.twins import ANSWER_FILL, _freeze

    wb = Workbook()
    ws = wb.active
    ws.title = "Scope"
    ws["A1"], ws["B1"], ws["C1"] = "Ref", "Question", "Response"
    rows = [
        ("1", "General capability questions for the vendor.", False),
        ("1.0.1", "Do you operate a regional delivery center?", True),
        ("1.0.2", "If yes to 1.0.1, describe the delivery center staffing model.", True),
    ]
    for i, (ref, q, filled) in enumerate(rows, start=2):
        ws[f"A{i}"], ws[f"B{i}"] = ref, q
        if filled:
            ws[f"C{i}"] = ""
            ws[f"C{i}"].fill = ANSWER_FILL
    _freeze(wb, tmp_path / "hdr.xlsx")

    parsed = parse_workbook(tmp_path / "hdr.xlsx")
    by_ref = {s.get("ref_id"): s for s in parsed.slots}
    header = by_ref["1"]
    assert header["is_header"] is True
    assert header["response_shape"] == "none"
    assert by_ref["1.0.1"]["parent"] == header["slot_id"]
    assert by_ref["1.0.1"]["response_shape"] == "boolean"
    # Symmetric gating: the child names its gater; the gater lists the child.
    assert by_ref["1.0.2"]["gating"]["gated_by"] == "1.0.1"
    assert by_ref["1.0.1"]["gating"]["gates"] == [by_ref["1.0.2"]["slot_id"]]


def test_record_rows_from_labeled_columns(tmp_path):
    """Rule 5b: on a NUMBERED questionnaire sheet (the hosp-erp record-
    section shape — a ref column is what distinguishes record rows from
    a grid sheet), >=2 answer cells covered by a valid label row become
    one record slot per row with typed fields."""
    from tests.fixtures.twins import ANSWER_FILL, _freeze

    wb = Workbook()
    ws = wb.active
    ws.title = "Team"
    ws["A1"], ws["B1"] = "Ref", "Question"
    ws["C1"], ws["D1"], ws["E1"] = "Name", "Role", "Hourly rate"
    rows = [
        ("5.0.1", "Provide the proposed team members for this engagement."),
        ("5.0.2", "Provide the proposed subcontractor personnel, if any."),
    ]
    for i, (ref, question) in enumerate(rows, start=2):
        ws[f"A{i}"], ws[f"B{i}"] = ref, question
        for col in ("C", "D", "E"):
            ws[f"{col}{i}"] = ""
            ws[f"{col}{i}"].fill = ANSWER_FILL
    _freeze(wb, tmp_path / "rec.xlsx")

    parsed = parse_workbook(tmp_path / "rec.xlsx")
    records = [s for s in parsed.slots if s["response_shape"] == "record"]
    assert len(records) == 2
    assert records[0]["ref_id"] == "5.0.1"
    assert records[0]["question_text"].startswith("Provide the proposed team")
    assert [(f["key"], f["type"]) for f in records[0]["response_fields"]] == [
        ("name", "text"), ("role", "text"), ("hourly_rate", "currency"),
    ]
