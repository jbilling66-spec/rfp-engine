"""Drafting test harness: the P6 chain extended one stage (run_0005).

run_drafting_package chains intake -> KB -> research -> strategy ->
Gate 1 -> [canonical plant] -> planning -> Gate 2 (dispositions) ->
drafting. The canonical card is planted BEFORE planning through the
real store seam, so P6's own mapper ranks it into kb_hits and the RAG
ban stays intact — the plan selected it, drafting just opened it. The
plant is asserted into the plan at the fixture (a ranking shift fails
here, loudly, not in an acceptance test).

make_drafter_script is the derive-wire-from-prompt fake (house pattern
#1): it reads back the SLOT directive lines, kb_card frames, brief
frame (REUSING tests/planning/fixtures/plans.py's regexes — the same
renderer serves planning and drafting, so drift fails both suites
together), lead frame, and the canonical/flag demands, and emits a wire
that mirrors terminology, threads themes, reproduces canonical bodies,
and flags where demanded — so fixture, prompt, frame, and script cannot
drift. Misbehavior kwargs plant each drop-and-report row.

The planning block is deliberately duplicated from plans.py's
run_planning_package rather than parameterizing the frozen fixture
(freeze discipline): the plant hook must sit between corpus ingest and
path mapping, a seam the P6 harness rightly never exposed.
"""

import hashlib
import json
import re
from pathlib import Path

from engine.drafting import run_drafting
from engine.drafting.compose import VOICE_DEFAULT
from engine.kb import KBStore
from engine.llm import FakeCaller, TracedCaller, effective_config
from engine.planning import approve_gate2, run_planning
from engine.runlog import RunLogger
from engine.version import engine_version
from tests.planning.fixtures.plans import (
    _TERMS_RX,
    _THEME_LINE_RX,
    ACTOR,
    BRIEF_RX,
    FIXTURES,
    GATE2_AT,
    LEAD_RX,
    WORKBOOKS,
    make_architect_script,
    planning_extras,
)
from tests.strategy.fixtures.strategies import run_strategy_package

DRAFT_AT = GATE2_AT  # drafting stamps no time of its own; gates own clocks

CANON_ID = "kb_canon000001"
CANONICAL_BODY = (
    "Our data migration factory runs trial loads, reconciliation gates, and "
    "a rehearsed cutover runbook for patient and financial records."
)

# Frame/directive regexes — single source for the derive-wire script AND
# the drift assertions in test_draft_frames (plans.py precedent).
VOICE_RX = re.compile(r'<voice_spec label="firm">\n(.*?)\n</voice_spec>', re.S)
KB_CARD_RX = re.compile(
    r'<kb_card kb_id="([^"]+)" title="([^"]*)" label="firm">\n(.*?)\n</kb_card>',
    re.S,
)
SLOT_RX = re.compile(r"^SLOT (\S+) \| ref (\S+)$", re.M)
SECTION_RX = re.compile(r"^SECTION: (.+)$", re.M)
CANON_RX = re.compile(r"^CANONICAL (\S+): ", re.M)
FLAG_SLOTS_RX = re.compile(r"^FLAGGED DRAFTING for slots: (.+?) — ", re.M)
FLAG_SECTION_MARK = "FLAGGED DRAFTING for this section"

ANSWERED_TEXT = ("We hold a current SOC 2 Type II attestation for hosted "
                 "operations.")
FLAG_NOTE = "Best-effort draft; flag every novel claim."


def drafting_extras(voice_path: Path | None = None) -> dict:
    """Planning extras + the voice spec digest (B31(13)): two drafting
    runs are comparable only when the voice content matches."""
    extras = planning_extras()
    extras["drafting"] = {"voice_spec_sha256": hashlib.sha256(
        Path(voice_path or VOICE_DEFAULT).read_bytes()).hexdigest()}
    return extras


def plant_canonical_card(store: KBStore) -> str:
    """The canonical-block fixture card, written through the REAL seam
    (tests/kb inline-card precedent). Catalog text is saturated with the
    gapcase 1.0.1 question tokens so the mapper ranks it into
    1-delivery-approach's kb_hits above the grounding floor."""
    store.write_card(
        {"kb_id": CANON_ID, "layer": "corpus",
         "title": "Data migration approach for patient and financial records",
         "summary": ("Approved boilerplate: describe the data migration "
                     "approach for patient and financial records."),
         "section_types": ["data_migration"],
         "canonical_block": True, "use_restriction": False, "outcome": "won"},
        CANONICAL_BODY,
        {"source_pursuit": "pur_synth", "source_client":
         "Meridian Health Partners", "date": "2025-01-01",
         "ingested_by": "ingestion_agent"},
        {},
    )
    return CANON_ID


def default_dispose(pursuit, package_id: str) -> list[dict]:
    """Gapcase: the content-test mix (answered + draft_flagged). Every
    other package: dispose every open gap draft_flagged so every section
    drafts regardless of corpus drift."""
    if package_id == "gapcase":
        pid = pursuit.pursuit_id
        return [
            {"section_id": "2-special-requirements",
             "gap_id": f"gap_{pid}_plan_01", "action": "answered",
             "answer": ANSWERED_TEXT},
            {"section_id": "2-special-requirements",
             "gap_id": f"gap_{pid}_plan_02", "action": "draft_flagged",
             "note": FLAG_NOTE},
        ]
    plan = pursuit.read_artifact("plan.json")
    return [
        {"section_id": s["section_id"], "gap_id": g["gap_id"],
         "action": "draft_flagged", "note": FLAG_NOTE}
        for s in plan["sections"] for g in s.get("gaps", [])
        if g.get("status") == "open"
    ]


def run_drafting_package(tmp_root, *, package_id="gapcase", script=None,
                         fake=None, dispose=None, plant_canonical=True,
                         workbook=None, voice_path=None, ceiling=None):
    """Full chain to a drafted pursuit (runs 0001-0005)."""
    pursuit, _ = run_strategy_package(tmp_root, package_id=package_id)
    store = KBStore(tmp_root / "kb")
    if plant_canonical:
        plant_canonical_card(store)

    # planning run 0004 (see module docstring on the duplication)
    log = RunLogger(pursuit.root, pursuit.new_run_id(), pursuit.pursuit_id)
    caller = TracedCaller(FakeCaller(make_architect_script()), log)
    cfg = effective_config(extra=drafting_extras(voice_path))
    log.run_start(mode="dry_run", engine_version=engine_version(), config=cfg,
                  kb_snapshot=store.snapshot(),
                  research_mode=cfg["research_mode"])
    if workbook is None:
        name = WORKBOOKS.get(package_id)
        workbook = FIXTURES / name if name else None
    report = run_planning(pursuit, caller, log, store, workbook=workbook)
    assert report.status == "complete", f"planning refused: {report}"
    if dispose is None:
        dispose = default_dispose(pursuit, package_id)
    approve_gate2(pursuit, log,
                  decision="approved_with_edits" if dispose else "approved",
                  actor=ACTOR, at=GATE2_AT,
                  edits={"dispose": dispose} if dispose else None)
    log.run_end(status="completed")

    if plant_canonical and package_id == "gapcase":
        frozen = pursuit.read_artifact("plan.frozen.json")
        hits = {h["kb_id"] for s in frozen["sections"]
                for h in s.get("kb_hits", [])}
        assert CANON_ID in hits, (
            "canonical plant failed to rank into kb_hits — fix the fixture "
            "card's catalog text, not the acceptance tests")

    return run_drafting_run(tmp_root, pursuit, script=script, fake=fake,
                            voice_path=voice_path, ceiling=ceiling)


def run_drafting_run(tmp_root, pursuit, *, script=None, fake=None,
                     voice_path=None, ceiling=None):
    """The drafting run alone, over an existing workspace — resume and
    refusal tests re-enter here. No try/except: a scripted kill or a
    fired ceiling propagates to the test, leaving the honest no-footer
    run behind."""
    store = KBStore(tmp_root / "kb")
    log = RunLogger(pursuit.root, pursuit.new_run_id(), pursuit.pursuit_id)
    caller = TracedCaller(fake or FakeCaller(script or make_drafter_script()),
                          log, **({"ceiling_usd": ceiling} if ceiling else {}))
    cfg = effective_config(extra=drafting_extras(voice_path))
    log.run_start(mode="dry_run", engine_version=engine_version(), config=cfg,
                  kb_snapshot=store.snapshot(),
                  research_mode=cfg["research_mode"])
    kwargs = {"voice_path": Path(voice_path)} if voice_path else {}
    report = run_drafting(pursuit, caller, log, store, **kwargs)
    log.run_end(status="completed")
    return pursuit, report


class SpyCaller(FakeCaller):
    """FakeCaller that records every call's (agent, tier, prompt, system)
    — the v1 cacheable-prefix assertion needs the system string."""

    def __init__(self, script):
        super().__init__(script)
        self.calls = []

    def call_for(self, agent, *, tier, prompt, system=""):
        self.calls.append({"agent": agent, "tier": tier, "prompt": prompt,
                           "system": system})
        return super().call_for(agent, tier=tier, prompt=prompt,
                                system=system)


def make_drafter_script(*, plant_cite_unopened=False, plant_missing_slot=False,
                        plant_duplicate_slot=False,
                        plant_unrequested_slot=False,
                        plant_alter_canonical=False, plant_omit_flag=False,
                        check_garbage=False, check_fix=None,
                        fail_on_section=None):
    """The well-behaved drafter fake, with plantable misbehaviors —
    everything it emits is read back from the prompt the stage actually
    built."""

    def _draft(prompt: str) -> str:
        slot_ids = [m.group(1) for m in SLOT_RX.finditer(prompt)]
        cards = [(m.group(1), m.group(3))
                 for m in KB_CARD_RX.finditer(prompt)]
        card_ids = [kb_id for kb_id, _ in cards]
        brief_m = BRIEF_RX.search(prompt)
        brief = brief_m.group(1) if brief_m else ""
        terms_m = _TERMS_RX.search(brief)
        terms = ([t.strip() for t in terms_m.group(1).split(";")]
                 if terms_m else [])
        themes = [m.group(1) for m in _THEME_LINE_RX.finditer(brief)]
        lead_m = LEAD_RX.search(prompt)
        lead_tail = (lead_m.group(1).splitlines()[-1] if lead_m else "")
        canon_ids = CANON_RX.findall(prompt)
        canon_bodies = {kb_id: body for kb_id, body in cards
                        if kb_id in canon_ids}
        flags_m = FLAG_SLOTS_RX.search(prompt)
        flagged = ([s.strip() for s in flags_m.group(1).split(",")]
                   if flags_m else [])
        flag_all = FLAG_SECTION_MARK in prompt

        def prose_for(slot_id: str, first: bool) -> str:
            parts = [f"Our engagement plan answers {slot_id} with delivery "
                     "commitments grounded in prior work."]
            parts.extend(f"We tailor it to your {term}." for term in terms)
            parts.extend(f"{theme}." for theme in themes)
            if lead_tail:
                parts.append(f"Per your direction: {lead_tail}")
            if first:
                for kb_id in canon_ids:
                    body = canon_bodies.get(kb_id, "")
                    if plant_alter_canonical:
                        body = body.replace("rehearsed", "improvised")
                    parts.append(body)
            if (flag_all or slot_id in flagged) and not plant_omit_flag:
                parts.append("[proposed approach] We propose a phased "
                             "accelerator novel to this engagement.")
            return " ".join(p for p in parts if p)

        cited = list(card_ids)
        if plant_cite_unopened:
            cited = cited + ["kb_unopened99"]
        if slot_ids:  # Path A arm
            entries = [{"slot_id": sid, "prose": prose_for(sid, i == 0),
                        "kb_ids": cited}
                       for i, sid in enumerate(slot_ids)]
            if plant_missing_slot and entries:
                entries.pop()
            if plant_duplicate_slot and entries:
                entries.append(dict(entries[0]))
            if plant_unrequested_slot:
                entries.append({"slot_id": "sl_ghost",
                                "prose": "never asked", "kb_ids": []})
            return json.dumps({"answers": entries})
        return json.dumps({"prose": prose_for("this section", True),
                           "kb_ids": cited})

    def _check(prompt: str) -> str:
        if check_garbage:
            return "not json {{{"
        if check_fix is not None:
            return (check_fix if isinstance(check_fix, str)
                    else json.dumps(check_fix))
        return json.dumps({"verdict": "pass"})

    def _dispatch(prompt: str) -> str:
        _dispatch.prompts.append(prompt)
        if fail_on_section and not prompt.startswith("Task: check."):
            section_m = SECTION_RX.search(prompt)
            if section_m and section_m.group(1) == fail_on_section:
                raise RuntimeError(
                    f"killed mid-run at {fail_on_section} (test)")
        return (_check(prompt) if prompt.startswith("Task: check.")
                else _draft(prompt))

    _dispatch.prompts = []
    return {"section_drafter": _dispatch}


def make_workbook_variant(tmp_path, *, instructions_extra: tuple[str, ...]
                          ) -> Path:
    """A gapcase copy with extra Instructions lines, saved to tmp — the
    REAL parser stamps the resulting global constraints on every slot;
    no new committed binary enters the tripwire sweep."""
    from openpyxl import load_workbook

    wb = load_workbook(FIXTURES / WORKBOOKS["gapcase"])
    ws = wb["Instructions"]
    row = ws.max_row
    for i, line in enumerate(instructions_extra, start=row + 1):
        ws[f"A{i}"] = line
    path = Path(tmp_path) / "gapcase-variant.xlsx"
    wb.save(path)
    return path


def read_draft(pursuit) -> dict:
    return pursuit.read_artifact("drafts/draft.json")


def section_by_id(envelope: dict, section_id: str) -> dict:
    return next(s for s in envelope["sections"]
                if s["section_id"] == section_id)


def answer_by_ref(section: dict, ref_id: str) -> dict:
    return next(a for a in section.get("answers", [])
                if a.get("ref_id") == ref_id)
