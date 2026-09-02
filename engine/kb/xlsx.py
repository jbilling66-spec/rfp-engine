"""The KB workbook: export for review, import as proposals.

Columns are DERIVED FROM THE SCHEMA, never hand-listed — v1's lesson,
where a hand-kept column list is the registry that drifts. The export is
what a steward edits and the import is what reads it back, so the two
halves share one column definition and cannot disagree about what a
column means.

Round-trip discipline, the transferable rule from v1: **diff PARSED
MODELS, never serialized text**. A textual comparison reports every card
as edited the first time anyone opens the sheet, and an import that
starts with noise a steward must read past is how review becomes
rubber-stamping.

Import produces PROPOSALS (S4). Nothing here writes a card.
"""

import json
from datetime import datetime, timezone
from pathlib import Path
from engine.structure.zipguard import ZipGuardError, check_office_zip

ROOT = Path(__file__).resolve().parents[2]
SCHEMA_PATH = ROOT / "schemas" / "kb-card.schema.json"
SHEET = "cards"
LIST_SEP = "; "
MAX_ROWS = 5000
HEADER_SCAN_ROWS = 6

# Never importable. kb_id identifies the row; the rest are governance or
# derived state, and a spreadsheet is not where those get decided.
#   layer / sensitivity / use_restriction / legal_hold / canonical_block
#     — governance: who may see this and whether it is boilerplate.
#   edit_survival / version — derived by the engine; a typed-over signal
#     would be a measurement nobody took.
#   grain / canonical_doc_id / content_origin / figure_class /
#   extraction_status — engine-derived canonical-model facts (WP13); a
#     typed-over value would be a measurement nobody took.
READ_ONLY = frozenset({
    "kb_id", "layer", "sensitivity", "use_restriction", "legal_hold",
    "canonical_block", "edit_survival", "version",
    "grain", "canonical_doc_id", "content_origin", "figure_class",
    "extraction_status",
})

# Structured sub-objects a flat sheet cannot carry honestly; doc_path is
# navigation-only (WP13 KB10) and never a steward-editable fact.
SKIP = frozenset({"anonymization", "provenance", "content_path",
                  "identity", "chunk_span", "doc_path"})


def _schema() -> dict:
    return json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))


def columns() -> list[tuple[str, str]]:
    """(name, kind) derived from the card schema. kind drives coercion:
    list | boolean | integer | number | string."""
    props = _schema()["properties"]
    out = []
    for name, spec in props.items():
        if name in SKIP:
            continue
        declared = spec.get("type")
        if declared == "array":
            kind = "list"
        elif declared == "boolean":
            kind = "boolean"
        elif declared == "integer":
            kind = "integer"
        elif declared == "number":
            kind = "number"
        else:
            kind = "string"
        out.append((name, kind))
    return out


class WorkbookError(ValueError):
    """The workbook could not be read at all — refused before any row."""


def _cell_value(front: dict, name: str, kind: str):
    value = front.get(name)
    if value is None:
        return None
    if kind == "list":
        joined = LIST_SEP.join(str(v) for v in value)
        for item in value:
            if LIST_SEP.strip() in str(item):
                # v1 split comma-joined lists back apart on import and
                # silently lost data. Refusing loudly beats a round trip
                # that quietly changes the content.
                raise WorkbookError(
                    f"{front.get('kb_id')}: {name} value {item!r} contains "
                    f"{LIST_SEP.strip()!r}, which is the list separator — "
                    f"export refuses rather than produce a sheet that "
                    f"cannot round-trip")
        return joined
    return value


def export_cards(store, path: Path) -> Path:
    """Write the review workbook. Typed cells (not str()-coerced), so a
    boolean reads as a boolean and a number is not quoted."""
    from openpyxl import Workbook

    book = Workbook()
    sheet = book.active
    sheet.title = SHEET
    cols = columns()
    sheet.append([name for name, _ in cols])
    for front in sorted(store.list_cards(), key=lambda c: c.get("kb_id", "")):
        sheet.append([_cell_value(front, name, kind) for name, kind in cols])
    _finalize(book, path)
    return path


def _finalize(book, path: Path) -> Path:
    """openpyxl stamps dcterms:modified from the wall clock inside
    save(), so two identical exports differ by bytes across a second
    boundary (v1's intermittent-flake lesson). Pinning it keeps an
    unchanged pack producing an unchanged file."""
    book.properties.created = datetime(2026, 1, 1, tzinfo=timezone.utc).replace(
        tzinfo=None)
    book.properties.modified = book.properties.created
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    return path


# ------------------------------------------------------------- import

def _coerce(raw, kind: str, address: str, name: str):
    """Excel retypes cells; this puts them back. Returns (value, error)."""
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None, None
    if isinstance(raw, str) and raw.lstrip().startswith("="):
        return None, (address, "formula_cell",
                      f"{name} holds a formula — paste values instead, so "
                      f"the sheet carries what you meant rather than how "
                      f"it was computed")
    if kind == "list":
        return [part.strip() for part in str(raw).split(LIST_SEP.strip())
                if part.strip()], None
    if kind == "boolean":
        text = str(raw).strip().lower()
        if text in ("true", "yes", "y", "1"):
            return True, None
        if text in ("false", "no", "n", "0"):
            return False, None
        return None, (address, "not_a_boolean",
                      f"{name} should be true or false, not {raw!r}")
    if kind in ("integer", "number"):
        try:
            return (int(raw) if kind == "integer" else float(raw)), None
        except (TypeError, ValueError):
            return None, (address, "not_a_number",
                          f"{name} should be a number, not {raw!r}")
    if isinstance(raw, datetime):
        return raw.date().isoformat(), None
    return str(raw).strip(), None


def _find_header(sheet, known: set[str]):
    """Found, not assumed: a banner row above the header is normal in a
    workbook someone has been working in."""
    for row in sheet.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS):
        names = [str(c.value).strip().lower() if c.value is not None else ""
                 for c in row]
        if sum(1 for n in names if n in known) >= 3:
            return row[0].row, names
    return None, []


def read_workbook(path: Path) -> tuple[list[dict], list[tuple], list[str]]:
    """Parse to (rows, errors, warnings). Rows carry `_address` so every
    error can name the cell a human should click."""
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell

    try:
        # data_only=False keeps formulas visible AS formulas: with
        # data_only=True a workbook saved by a non-Excel tool has no
        # cached value, so "=CONCAT(...)" reads as an empty required
        # cell while the user is looking at visible text.
        try:
            check_office_zip(Path(path))  # P0-8: the container first
        except ZipGuardError as exc:
            raise WorkbookError(str(exc)) from exc
        book = load_workbook(Path(path), data_only=False)
    except Exception as exc:  # noqa: BLE001 — any openpyxl failure
        raise WorkbookError(
            f"could not read the workbook ({type(exc).__name__}) — is it a "
            f"KB export?") from exc
    if SHEET not in book.sheetnames:
        raise WorkbookError(
            f"no {SHEET!r} sheet — is this a KB export? found: "
            f"{book.sheetnames}")
    sheet = book[SHEET]

    cols = columns()
    kinds = dict(cols)
    known = {name.lower() for name, _ in cols}

    merged = [str(r) for r in sheet.merged_cells.ranges]
    if merged:
        raise WorkbookError(
            f"merged cells are not supported — unmerge {', '.join(merged)}. "
            f"(A merged range reads as empty in every cell but the first, "
            f"which would look like missing required values.)")

    header_row, names = _find_header(sheet, known)
    if header_row is None:
        raise WorkbookError(
            f"no header row found in the first {HEADER_SCAN_ROWS} rows")

    seen, duplicates = {}, []
    for index, name in enumerate(names):
        if name in known:
            if name in seen:
                duplicates.append(name)
            else:
                seen[name] = index
    if duplicates:
        raise WorkbookError(
            f"duplicate column(s) {sorted(set(duplicates))} — the sheet is "
            f"ambiguous about which one to read")
    missing = sorted(known - set(seen))
    if missing:
        raise WorkbookError(
            f"missing column(s) {missing} — refusing before reading any row")

    warnings = []
    unknown = sorted({n for n in names if n and n not in known})
    if unknown:
        warnings.append(f"ignored unknown column(s): {', '.join(unknown)}")

    rows, errors = [], []
    body = list(sheet.iter_rows(min_row=header_row + 1))
    if len(body) > MAX_ROWS:
        raise WorkbookError(
            f"{len(body)} rows exceeds the {MAX_ROWS}-row cap")
    for row in body:
        if all(c.value is None or (isinstance(c.value, MergedCell))
               or str(c.value).strip() == "" for c in row):
            continue
        parsed = {"_address": f"{SHEET}!A{row[0].row}", "_row": row[0].row}
        for name, index in seen.items():
            cell = row[index] if index < len(row) else None
            address = (f"{SHEET}!{cell.coordinate}" if cell is not None
                       else parsed["_address"])
            value, error = _coerce(cell.value if cell else None,
                                   kinds[name], address, name)
            if error:
                errors.append(error)
            elif value is not None:
                parsed[name] = value
        rows.append(parsed)
    return rows, errors, warnings


def plan_import(store, path: Path) -> dict:
    """Read the sheet and decide what it would change. Writes NOTHING.

    ALL-OR-NOTHING (v1's rule, the owner's before that): any error and no
    proposal is opened. A partial import leaves a steward guessing which
    of their forty edits landed, and the pack in a state nobody chose.
    Every error is listed — never a sample — with the cell to click.
    """
    rows, errors, warnings = read_workbook(path)

    known_ids = {c["kb_id"] for c in store.list_cards()}
    seen_rows: dict[str, int] = {}
    changes: list[dict] = []

    for row in rows:
        kb_id = row.get("kb_id")
        if not kb_id:
            errors.append((row["_address"], "empty_required_field",
                           "kb_id is required and this cell is empty"))
            continue
        if kb_id in seen_rows:
            errors.append((row["_address"], "duplicate_kb_id",
                           f"{kb_id} also appears at row {seen_rows[kb_id]}"))
            continue
        seen_rows[kb_id] = row["_row"]

        if kb_id not in known_ids:
            # Minting belongs to the curation screen, where provenance
            # and anonymization are asked for. A blank row in a
            # spreadsheet cannot supply either.
            errors.append((row["_address"], "unknown_kb_id",
                           f"{kb_id} is not in the knowledge base — new "
                           f"cards are created on the curation screen, "
                           f"which asks for provenance"))
            continue

        front, _body = store.read_card(kb_id)
        diff = {}
        for name, _kind in columns():
            if name in READ_ONLY:
                if name in row and row[name] != front.get(name):
                    errors.append((
                        row["_address"], "locked_field",
                        f"{name} is not a cell edit — it is a governance "
                        f"decision (or an engine-derived value), so the "
                        f"sheet shows it and will not import a change"))
                continue
            if name in row and row[name] != front.get(name):
                # The diff is on PARSED values, so formatting noise never
                # presents as an edit.
                diff[name] = {"before": front.get(name), "after": row[name]}
        if diff:
            changes.append({"kb_id": kb_id, "diff": diff})

    return {"status": "refused" if errors else "ok",
            "errors": [{"cell": cell, "code": code, "message": message}
                       for cell, code, message in errors],
            "warnings": warnings,
            "changes": changes,
            "unchanged": len(rows) - len(changes) - len(errors)}


def submit_import(store, path: Path, *, operator: str, at: str) -> dict:
    """Plan, then open one proposal per changed card (S4). Refuses whole
    if the plan found any error."""
    from engine.flywheel.proposals import ProposalStore

    plan = plan_import(store, path)
    if plan["status"] != "ok":
        plan["proposals"] = []
        return plan

    proposals = ProposalStore(store.root)
    opened = []
    for change in plan["changes"]:
        proposal = proposals.open(
            source={"door": "xlsx_import", "operator": operator},
            target="corpus", kind="update_card", at=at,
            kb_id=change["kb_id"], diff=change["diff"],
            note=f"Edited in the KB workbook by {operator}.")
        opened.append(proposal["proposal_id"])
    plan["proposals"] = opened
    plan["message"] = (
        f"{len(opened)} change proposal(s) created; {plan['unchanged']} "
        f"card(s) identical — nothing happened to them")
    return plan
