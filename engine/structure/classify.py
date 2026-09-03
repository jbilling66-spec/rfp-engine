"""Layer 3 — classification: facts + conventions -> TargetSlot dicts.

Rules, in order (v1 oracle):
1. Formula cells are FACTS, never slots (cross-sheet titles, =SUM totals).
2. Column-header/label rows are convention rows — never slots. A
   ref-carrying label row can be BOTH a header slot and a label source.
   Label rows lose validity at the next header slot (subsection reset).
3. HEADER SLOTS are rows whose ref value is shallower than the sheet's
   leaf depth. Ref-depth is THE signal — decorated no-ref rows are
   furniture.
4. LEAF rows: any leaf-depth-ref row; or, without a ref, a question-text
   row with an answer cell. Shape by trigger tables (boolean first),
   criterion forward-filled from merged ranges, appendix read from
   answer-cell directives then question-text signals.
5. RECORD rows: >=2 answer cells covered by a VALID label row. Grids
   (ref-less sheets whose data rows carry >=2 answer cells or labeled
   values) collapse to one template_fill slot per grid; formula cells
   excluded.

Slots are plain dicts shaped exactly per the frozen
schemas/target-slot.schema.json; absent fields are OMITTED, never
nulled (the v2 writers-omit rule — v1's null-bearing fixtures fail the
schema).
"""

import re

from openpyxl.utils import column_index_from_string, get_column_letter

from engine.contracts.slots import field_key
from engine.structure.conventions import (
    Conventions,
    SheetConventions,
    answer_cells,
    depth,
    merged_member_cols_by_row,
    row_ref,
)
from engine.structure.facts import SheetFacts

_REF_MENTION = re.compile(r"\b\d+\.\d+\.\d+\b")

# Boolean checked BEFORE numeric; openers only.
BOOLEAN_TRIGGERS = (
    "do ", "does ", "are ", "is ", "will ", "have ", "has ", "can ",
    "confirm",
)
NUMERIC_TRIGGERS = (
    "how many", "how much", "how long", "how frequently", "number of",
    "what percentage",
)
# Word-boundary matched — "county" must not trigger on "count".
_NUMERIC_LABEL = re.compile(r"\b(years|count|number|hours)\b|%")

# The ask wants WORDS back — vetoes numeric shape and typed columns
# (v1 precision pass: every term fires on a measured false positive).
_ASKS_FOR_PROSE = re.compile(
    r"\b(descri\w*|summar\w*|explain\w*|narrative|comments?|discuss\w*)\b",
    re.IGNORECASE,
)

FIELD_TYPE_TRIGGERS: tuple[tuple[str, str], ...] = (
    ("price", "currency"),
    ("cost", "currency"),
    ("rate", "currency"),
    ("fee", "currency"),
    ("%", "percent"),
    ("hours", "number"),
    ("duration", "number"),
    ("weeks", "number"),
    ("no.", "number"),
)

APPENDIX_DIRECTIVE = re.compile(
    r"do not insert here.*?(?:as part of|part of|within|in|as)\s+"
    r"(?:the\s+|an?\s+)?(.+?\bappendix)",
    re.IGNORECASE | re.DOTALL,
)
_APPENDIX_GENERIC = re.compile(
    r"do not insert here|separate attachment|separate appendix"
    r"|include as an appendix",
    re.IGNORECASE,
)

_GATED_BY = re.compile(r"if\s+yes\s+to\s+(\d+(?:\.\d+)+)", re.IGNORECASE)
_SENTENCE_SPLIT = re.compile(r"(?<=[.?!])\s+")
_SHORT_LABEL = 60


def _classify_shape(text: str) -> str:
    lowered = text.strip().lower()
    if any(lowered.startswith(t) for t in BOOLEAN_TRIGGERS):
        return "boolean"
    if _numeric_ask(lowered):
        return "numeric"
    if "?" not in text and len(text) <= _SHORT_LABEL:
        # Label-form field ("Years in business") — never sentence-form prose.
        if _NUMERIC_LABEL.search(lowered) and not lowered.startswith(
            ("describe", "explain", "summarize", "list", "provide", "state")
        ):
            return "numeric"
    return "prose"


def _numeric_ask(lowered: str) -> bool:
    """Prose veto + first-sentence-only (v1 precision pass): a question
    that wants a description is prose however many quantities it names;
    a trigger buried in a later sentence is an aside, not the ask."""
    if _ASKS_FOR_PROSE.search(lowered):
        return False
    first = _SENTENCE_SPLIT.split(lowered, maxsplit=1)[0]
    return any(t in first for t in NUMERIC_TRIGGERS)


def _field_type(label: str) -> str:
    lowered = label.lower()
    if _ASKS_FOR_PROSE.search(lowered):
        return "text"
    for term, ftype in FIELD_TYPE_TRIGGERS:
        if re.search(rf"(?<!\w){re.escape(term)}(?!\w)", lowered):
            return ftype
    return "text"


def _sub_questions(text: str) -> list[str]:
    sentences = [s.strip() for s in _SENTENCE_SPLIT.split(text) if s.strip()]
    interrogatives = [s for s in sentences if s.endswith("?")]
    return interrogatives if len(interrogatives) >= 2 else []


def _appendix(answer_text: str | None, question_text: str | None) -> str | None:
    """Answer-cell directives first (the real files' signal), then
    question-text directives."""
    for text in (answer_text, question_text):
        if not text:
            continue
        m = APPENDIX_DIRECTIVE.search(text)
        if m:
            return m.group(1).strip().title()
        if _APPENDIX_GENERIC.search(text):
            return "Appendix"
    return None


def _slot(**fields) -> dict:
    """Build a slot dict, omitting Nones, empty lists, and schema
    defaults (writers omit, never null)."""
    out = {}
    for key, value in fields.items():
        if value is None or value == [] or value == {}:
            continue
        if key == "is_header" and value is False:
            continue
        if key == "answer_location" and value == "inline":
            continue
        if key == "fill_type" and value == "authored":
            continue
        out[key] = value
    return out


def _leaf_path(sheet_name: str, ref: str | None, criterion: str | None) -> str:
    base = sheet_name.strip()
    if ref:
        return f"{base} > {ref}"
    if criterion:
        return f"{base} > {criterion}"
    return base


def _parent_for(last_by_depth: dict[int, str], d: int) -> str | None:
    shallower = [x for x in last_by_depth if x < d]
    return last_by_depth[max(shallower)] if shallower else None


def _current_parent(last_by_depth: dict[int, str]) -> str | None:
    return last_by_depth[max(last_by_depth)] if last_by_depth else None


def parse_sheet(
    sheet: SheetFacts, sc: SheetConventions, conv: Conventions, file_name: str,
    *, warnings: list[str] | None = None,
) -> list[dict]:
    """`warnings` (P26b-1, B112): the sink for every row that carried
    content and matched no rule — the P10-F16 class at the structure
    layer. Entries name the sheet, the row and the kind; never a cell's
    text. None keeps the pre-P26b behaviour (silent)."""
    slots: list[dict] = []
    if sc.kind == "instructions":
        return slots

    rows = sheet.rows()
    members = merged_member_cols_by_row(sheet)

    # Criterion forward-fill: merged vertical ranges in a text column.
    criterion_by_row: dict[int, str] = {}
    for rng in sheet.merged_ranges:
        m = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", rng)
        if not m or m.group(1) != m.group(3) or int(m.group(4)) <= int(m.group(2)):
            continue
        anchor = sheet.cells.get(f"{m.group(1)}{m.group(2)}")
        if anchor and anchor.text and not anchor.is_formula:
            col = column_index_from_string(m.group(1))
            if col not in (sc.ref_col, sc.question_col, sc.answer_col):
                for r in range(int(m.group(2)), int(m.group(4)) + 1):
                    criterion_by_row[r] = anchor.text.strip()

    last_header_by_depth: dict[int, str] = {}
    active_label_row: int | None = None
    labeled_cols = {c for cols in sc.label_rows.values() for c, _ in cols}

    for row_num, row_facts in sorted(rows.items()):
        answers = answer_cells(row_facts, sc, conv, row_num, members.get(row_num))

        # Rule 2: label rows are convention rows.
        if row_num in sc.label_rows:
            active_label_row = row_num
            if row_ref(row_facts, sc.ref_col) is None:
                continue
            # A ref-carrying label row falls through: header slot below.

        # Furniture above the first label row (titles, banners) — RECORDED
        # when the row carries a ref, i.e. looks like a question the
        # buyer placed above the header (P1-23); bare titles stay silent.
        if sc.first_label_row is not None and row_num < sc.first_label_row:
            if warnings is not None and row_ref(row_facts, sc.ref_col) is not None:
                warnings.append(f"{sheet.name.strip()}!row {row_num}: row above "
                                "the first label row skipped (carries a ref)")
            continue

        ref = row_ref(row_facts, sc.ref_col)
        # P1-24: a formula cell counts as text ONLY when it carries a
        # cached value (its `text` is then that value); answer-side rules
        # keep treating it as a formula.
        text_cells = [f for f in row_facts
                      if f.text and (not f.is_formula or f.cached_text is not None)]
        question = next(
            (f.text.strip() for f in text_cells if f.col == sc.question_col), None
        )
        answer_text = next(
            (f.text.strip() for f in text_cells
             if f.col == sc.answer_col and not f.is_formula), None
        )
        slot_id = f"slot_{sheet.index:02d}_r{row_num:03d}"

        # Rule 3: header slots — ref shallower than leaf depth, no answers.
        if ref and depth(ref) < sc.leaf_depth and not answers:
            slots.append(_slot(
                slot_id=slot_id,
                ref_id=ref,
                source_mode="client_provided",
                source_locator={"file": file_name, "sheet": sheet.name},
                path=f"{sheet.name.strip()} > {ref}",
                parent=_parent_for(last_header_by_depth, depth(ref)),
                is_header=True,
                question_text=question,
                response_shape="none",
            ))
            last_header_by_depth[depth(ref)] = slot_id
            for d in [d for d in last_header_by_depth if d > depth(ref)]:
                del last_header_by_depth[d]
            if row_num not in sc.label_rows:
                active_label_row = None  # subsection reset
            continue

        # Rule 5a: grid rows accumulate into one template_fill slot.
        if sc.kind == "grid":
            grid_cells = answers if len(answers) >= 2 else [
                f for f in row_facts
                if f.text and not f.is_formula and f.col in labeled_cols
            ]
            if len(grid_cells) >= 2 and (
                sc.first_label_row is None or row_num > sc.first_label_row
            ):
                _accumulate_grid(slots, sheet, sc, row_num, grid_cells, file_name)
            elif warnings is not None and any(f.text for f in row_facts):
                texts = [f for f in row_facts if f.text]
                kind = ("formula-only grid row skipped (totals are facts, "
                        "never targets)" if all(f.is_formula for f in texts)
                        else "grid row skipped (fewer than two labeled cells)")
                warnings.append(f"{sheet.name.strip()}!row {row_num}: {kind}")
            continue

        # Rule 5b: record rows — a VALID label row covers >=2 answer cells.
        if len(answers) >= 2 and active_label_row is not None:
            labels = sc.label_rows[active_label_row]
            covered = [(col, label) for col, label in labels
                       if any(f.col == col for f in answers)]
            if len(covered) >= 2:
                slots.append(_slot(
                    slot_id=slot_id,
                    ref_id=ref,
                    source_mode="client_provided",
                    source_locator={
                        "file": file_name, "sheet": sheet.name,
                        "cell": answers[0].cell,
                    },
                    path=(f"{sheet.name.strip()} > {ref}" if ref
                          else sheet.name.strip()),
                    parent=_current_parent(last_header_by_depth),
                    question_text=question,
                    response_shape="record",
                    response_fields=[
                        {"key": field_key(label), "label": label,
                         "type": _field_type(label),
                         "source_locator": {
                             "sheet": sheet.name,
                             "cell": f"{get_column_letter(col)}{row_num}",
                         }}
                        for col, label in covered
                    ],
                ))
                continue

        # Rule 4: leaf rows.
        is_leaf = bool(
            (ref and depth(ref) >= sc.leaf_depth)
            or (not ref and question and (answers or answer_text))
        )
        if is_leaf and (question or answers or answer_text):
            answer_cell = (
                answers[0].cell if answers
                else (f"{get_column_letter(sc.answer_col)}{row_num}"
                      if sc.answer_col is not None else None)
            )
            qtext = question or answer_text or ""
            appendix = _appendix(answer_text, question)
            gated = _GATED_BY.search(qtext)
            slots.append(_slot(
                slot_id=slot_id,
                ref_id=ref,
                source_mode="client_provided",
                source_locator={
                    "file": file_name, "sheet": sheet.name, "cell": answer_cell,
                },
                path=_leaf_path(sheet.name, ref, criterion_by_row.get(row_num)),
                parent=_current_parent(last_header_by_depth),
                question_text=question or answer_text,
                sub_questions=_sub_questions(qtext),
                eval_criterion=criterion_by_row.get(row_num),
                response_shape=_classify_shape(qtext),
                answer_location="appendix" if appendix else "inline",
                appendix_ref=appendix,
                gating={"gated_by": gated.group(1)} if gated else None,
                cross_refs=[r for r in _REF_MENTION.findall(qtext) if r != ref],
            ))
        elif warnings is not None and any(f.text for f in row_facts):
            # Everything else is furniture — but furniture that carries
            # content is a row the buyer wrote and the engine dropped,
            # so it is RECORDED (P1-23). Formula-only rows are the
            # P1-24 case and name themselves.
            texts = [f for f in row_facts if f.text]
            kind = ("formula-only row skipped (a formula question needs a "
                    "cached value)" if all(f.is_formula and f.cached_text is None
                                           for f in texts)
                    else "unclassified row dropped (no rule matched)")
            warnings.append(f"{sheet.name.strip()}!row {row_num}: {kind}")

    return slots


def _accumulate_grid(
    slots: list[dict],
    sheet: SheetFacts,
    sc: SheetConventions,
    row_num: int,
    grid_cells,
    file_name: str,
) -> None:
    """Ref-less grid rows collapse into ONE template_fill slot per grid,
    anchored at the grid's label row, fields from the REAL header labels."""
    candidates = [r for r in sc.label_rows if r < row_num]
    if not candidates:
        return
    header_row = max(candidates)
    labels = sc.label_rows[header_row]
    anchor_id = f"slot_{sheet.index:02d}_r{header_row:03d}"
    if any(s["slot_id"] == anchor_id for s in slots):
        return
    grid_cols = {f.col for f in grid_cells}
    title_fact = next(
        (f for f in sheet.rows().get(1, []) if f.text and not f.is_formula), None
    )
    slots.append(_slot(
        slot_id=anchor_id,
        source_mode="client_provided",
        source_locator={"file": file_name, "sheet": sheet.name},
        path=sheet.name.strip(),
        question_text=(
            f"Complete the buyer-defined grid: {title_fact.text.strip()}"
            if title_fact and header_row != 1
            else f"Complete the buyer-defined grid on {sheet.name.strip()}"
        ),
        response_shape="template_fill",
        fill_type="template_fill",
        response_fields=[
            {"key": field_key(label), "label": label, "type": _field_type(label),
             "source_locator": {"sheet": sheet.name,
                                "column": get_column_letter(col)}}
            for col, label in labels if col in grid_cols
        ],
    ))


def link_gating(slots: list[dict]) -> None:
    """Symmetric explicit gating links: "If yes to <ref>" sets gated_by;
    the gater gains the child in its gates list. (v1's subsection-prefix
    inference rule is deliberately not carried — no fixture exercises it;
    B28.)"""
    by_ref: dict[str, dict] = {}
    for slot in slots:
        ref = slot.get("ref_id")
        if ref and ref not in by_ref:
            by_ref[ref] = slot

    for slot in slots:
        gated_by = slot.get("gating", {}).get("gated_by")
        if gated_by and gated_by in by_ref:
            gater = by_ref[gated_by]
            gating = gater.setdefault("gating", {})
            gates = gating.setdefault("gates", [])
            if slot["slot_id"] not in gates:
                gates.append(slot["slot_id"])
