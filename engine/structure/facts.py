"""Layer 1 — workbook FACTS. Answers only "what is in the file": cell
text, fill signatures, formulas, merged ranges. It never decides what
anything MEANS — that is Layer 2 (conventions) and Layer 3 (classify).

Two hard rules from v1's real fixtures:
- Sheet names are BYTE-EXACT everywhere (trailing spaces are real; a
  stripped name breaks write-back).
- The workbook opens data_only=False + read_only=False — formulas must
  be visible as formulas, and merged ranges/fills must be reliable.
"""

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class CellFact:
    """One interesting cell: has text, a solid fill, or a formula."""

    cell: str  # A1 notation
    row: int
    col: int
    text: str | None  # stringified value; None for empty-but-filled cells
    fill: str | None  # fill signature or None
    is_formula: bool
    formula: str | None
    merged_range: str | None
    bold: bool


@dataclass
class SheetFacts:
    name: str  # BYTE-EXACT — never strip
    index: int  # physical tab position (0-based)
    cells: dict[str, CellFact] = field(default_factory=dict)
    merged_ranges: list[str] = field(default_factory=list)

    def rows(self) -> dict[int, list[CellFact]]:
        """Facts grouped by row, columns ordered."""
        out: dict[int, list[CellFact]] = {}
        for fact in self.cells.values():
            out.setdefault(fact.row, []).append(fact)
        for facts in out.values():
            facts.sort(key=lambda f: f.col)
        return out


@dataclass
class WorkbookFacts:
    file: str  # original basename
    sheets: list[SheetFacts] = field(default_factory=list)


def fill_signature(cell) -> str | None:
    """A fill's identity CLASS, not its resolved color: solid ARGB fills
    -> last-6 hex; theme/indexed fills -> tagged strings. Signature
    EQUALITY is all the convention learner needs."""
    f = cell.fill
    if f is None or f.patternType != "solid":
        return None
    color = f.fgColor
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        rgb = str(color.rgb)
        if rgb == "00000000":  # openpyxl default fill — not a fill
            return None
        return rgb[-6:].upper()
    if color.type == "theme":
        return f"theme:{color.theme}:{color.tint}"
    if color.type == "indexed":
        return f"indexed:{color.indexed}"
    return None


def collect_workbook_facts(path: Path) -> WorkbookFacts:
    """Read every interesting cell (text, fill, or formula) with
    byte-exact sheet names. data_only=False so formulas stay formulas."""
    wb = load_workbook(path, data_only=False, read_only=False)
    facts = WorkbookFacts(file=path.name)
    for index, ws in enumerate(wb.worksheets):
        sheet = SheetFacts(
            name=ws.title,
            index=index,
            merged_ranges=sorted(str(r) for r in ws.merged_cells.ranges),
        )
        by_cell_range: dict[str, str] = {}
        for rng in ws.merged_cells.ranges:
            for row in ws[str(rng)]:
                for cell in row:
                    by_cell_range[cell.coordinate] = str(rng)
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                fill = fill_signature(cell)
                is_formula = isinstance(value, str) and value.startswith("=")
                if value is None and fill is None:
                    continue
                sheet.cells[cell.coordinate] = CellFact(
                    cell=cell.coordinate,
                    row=cell.row,
                    col=cell.column,
                    text=None if value is None else str(value),
                    fill=fill,
                    is_formula=is_formula,
                    formula=str(value) if is_formula else None,
                    merged_range=by_cell_range.get(cell.coordinate),
                    bold=bool(cell.font is not None and cell.font.bold),
                )
        facts.sheets.append(sheet)
    return facts
