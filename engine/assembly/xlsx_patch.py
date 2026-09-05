"""The buyer's workbook, byte-for-byte, plus the answers (P1-19, P26b-3;
the owner's call, B119 §1a).

openpyxl is a fine READER and a lossy WRITER: on load/save it blanks
every formula's cached value (the value P26b-1's parser relies on),
drops images when Pillow is absent, and rewrites every part it touches.
The register's remedy — assert the part inventory, refuse on drift —
would have refused most Excel-authored forms. So the write-back no
longer writes through openpyxl at all: the output zip is rebuilt from
the source member by member, in the source's own order with each
member's own timestamp and compression, and only the sheet parts that
carry an answered cell are patched — the one `<c>` becomes an inline
string, its style kept, everything around it untouched. Shared strings,
styles, charts, drawings, comments, data validations, defined names,
VBA, images and every cached value survive because nothing rewrote them.

Two other parts may change: `docProps/core.xml` (the buyer-owned
metadata rule, P3-15 — the firm as last modifier, generator strings
stripped) and `xl/calcChain.xml`, from which a written-over formula
cell is removed so Excel never repairs the file on open.

`assert_roundtrip` is the proof, run after every write: the part lists
are equal and in order, every unpatched member is byte-identical, and
on every patched sheet the cell model outside the intended coordinates
— formulas AND their cached values — equals the source's.
"""

import re
import zipfile
from pathlib import Path

from engine.assembly.hygiene import stamp_core_xml
from engine.contracts import ContractError

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_XML = "http://www.w3.org/XML/1998/namespace"
CORE_PART = "docProps/core.xml"
CALC_CHAIN = "xl/calcChain.xml"
_COORD = re.compile(r"^([A-Z]{1,3})([0-9]{1,7})$")


def _tag(local: str) -> str:
    return f"{{{NS_MAIN}}}{local}"


def split_coord(coord: str) -> tuple[int, int]:
    """'C12' -> (column 3, row 12); refuses anything else by name."""
    match = _COORD.match(coord)
    if not match:
        raise ContractError(f"{coord!r} is not a cell coordinate")
    letters, row = match.groups()
    column = 0
    for char in letters:
        column = column * 26 + (ord(char) - 64)
    return column, int(row)


def sheet_parts(zf: zipfile.ZipFile) -> dict[str, str]:
    """{sheet name: part name} through xl/workbook.xml and its rels —
    never by guessing sheetN from the order."""
    from lxml import etree

    workbook = etree.fromstring(zf.read("xl/workbook.xml"))
    rels = etree.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    targets = {rel.get("Id"): rel.get("Target") for rel in rels}
    out = {}
    for sheet in workbook.iter(_tag("sheet")):
        target = targets[sheet.get(f"{{{NS_REL}}}id")]
        part = target.lstrip("/") if target.startswith("/") else f"xl/{target}"
        out[sheet.get("name")] = part
    return out


def _sheet_index(zf: zipfile.ZipFile, sheet_name: str) -> int:
    """The 1-based position calcChain's `i` attribute uses."""
    from lxml import etree

    workbook = etree.fromstring(zf.read("xl/workbook.xml"))
    for index, sheet in enumerate(workbook.iter(_tag("sheet")), start=1):
        if sheet.get("name") == sheet_name:
            return index
    raise ContractError(f"sheet {sheet_name!r} is not in the workbook")


def patch_sheet(xml: bytes, writes: dict[str, str]) -> tuple[bytes, list[str]]:
    """Set each coordinate to an inline string inside the sheet XML.
    An existing cell keeps its style and loses its formula/value; a
    missing cell is inserted in column order; a missing row in row
    order. Returns the new part and the coordinates that carried a
    formula (for calcChain)."""
    from lxml import etree

    root = etree.fromstring(xml)
    sheet_data = root.find(_tag("sheetData"))
    if sheet_data is None:
        raise ContractError("the sheet part has no sheetData element")
    rows = {int(row.get("r")): row for row in sheet_data.findall(_tag("row"))}
    had_formula: list[str] = []
    for coord, value in writes.items():
        column, row_number = split_coord(coord)
        row = rows.get(row_number)
        if row is None:
            row = etree.Element(_tag("row"), r=str(row_number))
            later = [r for n, r in rows.items() if n > row_number]
            if later:
                min(later, key=lambda r: int(r.get("r"))).addprevious(row)
            else:
                sheet_data.append(row)
            rows[row_number] = row
        cells = row.findall(_tag("c"))
        cell = next((c for c in cells if c.get("r") == coord), None)
        if cell is None:
            cell = etree.Element(_tag("c"), r=coord)
            later = [c for c in cells if split_coord(c.get("r"))[0] > column]
            if later:
                later[0].addprevious(cell)
            else:
                row.append(cell)
            # `spans` is a hint over the row's cell range; a cell outside
            # it is legal, but the hint is dropped rather than left wrong
            if "spans" in row.attrib:
                del row.attrib["spans"]
        if cell.find(_tag("f")) is not None:
            had_formula.append(coord)
        style = cell.get("s")
        for child in list(cell):
            cell.remove(child)
        cell.attrib.clear()
        cell.set("r", coord)
        if style is not None:
            cell.set("s", style)
        cell.set("t", "inlineStr")
        text = etree.SubElement(etree.SubElement(cell, _tag("is")), _tag("t"))
        text.text = value
        text.set(f"{{{NS_XML}}}space", "preserve")
    return (etree.tostring(root, xml_declaration=True, encoding="UTF-8",
                           standalone=True), had_formula)


def _patch_calc_chain(xml: bytes, drop: set[tuple[int, str]]) -> bytes:
    from lxml import etree

    root = etree.fromstring(xml)
    for entry in list(root):
        if (int(entry.get("i", "1")), entry.get("r")) in drop:
            root.remove(entry)
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8",
                          standalone=True)


def write_cells(source: Path, output: Path,
                writes: dict[tuple[str, str], str], *, firm: dict,
                at: str) -> dict:
    """Rebuild `output` from `source`, patching only what the answers
    touch. writes: {(sheet name, coordinate): text}."""
    source, output = Path(source), Path(output)
    by_sheet: dict[str, dict[str, str]] = {}
    for (sheet, coord), value in writes.items():
        by_sheet.setdefault(sheet, {})[coord] = value
    patched: list[str] = []
    formula_cells: set[tuple[int, str]] = set()
    with zipfile.ZipFile(source) as src:
        parts = sheet_parts(src)
        missing = sorted(set(by_sheet) - set(parts))
        if missing:
            raise ContractError(
                f"write-back names sheet(s) {missing} the workbook does "
                f"not have — {sorted(parts)}")
        part_writes = {parts[sheet]: cells for sheet, cells in by_sheet.items()}
        sheet_of_part = {parts[sheet]: sheet for sheet in by_sheet}
        members = src.infolist()
        rebuilt: dict[str, bytes] = {}
        for info in members:
            data = src.read(info.filename)
            if info.filename in part_writes:
                data, had_formula = patch_sheet(data, part_writes[info.filename])
                patched.append(info.filename)
                index = _sheet_index(src, sheet_of_part[info.filename])
                formula_cells.update((index, coord) for coord in had_formula)
            elif info.filename == CORE_PART:
                data = stamp_core_xml(data, firm=firm, at=at)
                patched.append(info.filename)
            rebuilt[info.filename] = data
        if formula_cells and CALC_CHAIN in rebuilt:
            rebuilt[CALC_CHAIN] = _patch_calc_chain(rebuilt[CALC_CHAIN],
                                                    formula_cells)
            patched.append(CALC_CHAIN)
        output.parent.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(output, "w") as dst:
            for info in members:
                copy = zipfile.ZipInfo(info.filename, date_time=info.date_time)
                copy.compress_type = info.compress_type
                copy.external_attr = info.external_attr
                copy.create_system = info.create_system
                dst.writestr(copy, rebuilt[info.filename])
    return {"parts": len(members), "patched_parts": sorted(patched),
            "formula_cells_overwritten": sorted(
                coord for _, coord in formula_cells)}


def _cell_model(path: Path, sheet_name: str) -> tuple[dict, dict]:
    """{coord: formula-or-value} and {coord: cached value} — the two
    loads P1-24 taught the parser."""
    from openpyxl import load_workbook

    formulas = load_workbook(path)[sheet_name]
    cached = load_workbook(path, data_only=True)[sheet_name]
    model = {c.coordinate: c.value for row in formulas.iter_rows()
             for c in row if c.value is not None}
    values = {c.coordinate: c.value for row in cached.iter_rows()
              for c in row if c.value is not None}
    return model, values


def assert_roundtrip(source: Path, output: Path,
                     intended: set[tuple[str, str]]) -> None:
    """The proof the write-back stands behind (the docx twin's name):
    equal part lists in order, every unpatched member byte-identical,
    and on each patched sheet the cell model outside the intended
    coordinates — formulas and cached values — equal to the source's."""
    with zipfile.ZipFile(source) as src, zipfile.ZipFile(output) as out:
        if src.namelist() != out.namelist():
            raise ContractError(
                "xlsx write-back changed the part inventory — "
                f"source {len(src.namelist())} parts, output "
                f"{len(out.namelist())}; refusing to hand back a changed "
                "workbook")
        parts = sheet_parts(src)
        touched = {parts[sheet] for sheet, _ in intended}
        touched |= {CORE_PART, CALC_CHAIN}
        for name in src.namelist():
            if name in touched:
                continue
            if src.read(name) != out.read(name):
                raise ContractError(
                    f"xlsx write-back drifted part {name} outside every "
                    "intended cell — refusing to hand back a changed "
                    "workbook")
    for sheet in sorted({sheet for sheet, _ in intended}):
        coords = {coord for s, coord in intended if s == sheet}
        src_model, src_cached = _cell_model(source, sheet)
        out_model, out_cached = _cell_model(output, sheet)
        for model_src, model_out, what in ((src_model, out_model, "cell"),
                                           (src_cached, out_cached,
                                            "cached value")):
            for coord in set(model_src) | set(model_out):
                if coord in coords:
                    continue
                if model_src.get(coord) != model_out.get(coord):
                    raise ContractError(
                        f"xlsx write-back drifted {what} {sheet}!{coord} "
                        "outside every intended target")
