"""Write-back into the buyer's own file (B37/D19): the ONE tool that
touches buyer files (S5/T4) — and it touches a COPY under
exports/writeback/, never the inbox original.

The cell authority chain: slots.json (source_sha256-bound to the exact
buyer file) supplies the cells; the FROZEN plan supplies which slots a
section owns (T6: only plan-named cells); the envelope supplies the
drafted prose. Only `response_shape: prose` + `fill_type: authored`
slots with drafted prose are WRITTEN. Everything else is a RECORDED
refusal in WritebackFacts: `refused_shape` (boolean / numeric / record /
table / template_fill — the pricing grid — / none: shapes the drafter
structurally cannot emit, v1's batch-contract lesson), `refused_unnamed`
(the plan names no such cell), `empty_no_prose` (the slot stays
honestly empty for the human — verify.py's posture).

Preview-then-confirm (S7): `preview_writeback` computes the facts and
the cell-model diff without writing anything; `run_writeback` re-derives
the SAME facts (the server's decision, not the client's echo), writes
the copy, and records WritebackFacts + the artifact lines. Diffs are
cell-model, never bytes; formulas outside the written cells survive
because only the named cells are assigned (v1 EC-5).
"""

import hashlib
import shutil

from openpyxl import load_workbook

from engine.contracts import ContractError

WRITEBACK_DIR = "exports/writeback"
FACTS_NAME = "exports/writeback-facts.json"


def _writable(slot: dict) -> tuple[bool, str]:
    shape = slot.get("response_shape")
    if shape != "prose":
        return False, (f"response_shape {shape!r} never takes model prose "
                       "(B28(4) shape-skip: the drafter structurally "
                       "cannot emit it)")
    if slot.get("fill_type") not in (None, "authored"):
        return False, (f"fill_type {slot.get('fill_type')!r} is not the "
                       "authored lane")
    return True, ""


def _cells(pursuit) -> tuple[dict, dict, dict, dict]:
    frozen_plan = pursuit.read_artifact("plan.frozen.json")
    if frozen_plan.get("path") != "A_designated":
        raise ContractError(
            "write-back fills a buyer workbook, and a non-designated plan "
            "has no workbook cells BY CONSTRUCTION: firm_default fills the "
            "firm template (template_fill lane) and a mixed target set "
            "plans as A_designated (B74§3a) — this guard is an invariant, "
            "not a reachable lane for real pursuits (B77§2 D5)")
    container = pursuit.read_artifact(
        frozen_plan.get("slots_ref", "slots.json"))
    envelope = pursuit.read_artifact("drafts/draft.json")
    slots_by_id = {s["slot_id"]: s for s in container["slots"]}
    prose_by_slot = {a["slot_id"]: a
                     for e in envelope.get("sections", [])
                     for a in e.get("answers", [])}
    planned_slots = {sid for s in frozen_plan.get("sections", [])
                     for sid in s.get("slot_ids", [])}
    return container, slots_by_id, prose_by_slot, planned_slots


def compute_facts(pursuit, *, at: str, confirmed_by: str,
                  binding: dict | None = None) -> dict:
    """binding (P18/C4): one entry from declared_deliverables — names
    the file, digest, and facts path this record covers. None keeps the
    flat single-source behavior byte-compatible (legacy facts name,
    top-level container digest)."""
    container, slots_by_id, prose_by_slot, planned_slots = _cells(pursuit)
    envelope = pursuit.read_artifact("drafts/draft.json")
    wanted = (binding["source_sha256"] if binding
              else container["source_sha256"])
    source = None
    for candidate in sorted((pursuit.root / "inbox").glob("*.xlsx")):
        digest = hashlib.sha256(candidate.read_bytes()).hexdigest()
        if digest == wanted:
            source = candidate
            break
    if source is None:
        raise ContractError(
            "no inbox workbook matches the slots container's "
            "source_sha256 — write-back only fills the EXACT file the "
            "slots were parsed from")
    cells = []
    for slot_id, slot in sorted(slots_by_id.items()):
        locator = slot.get("source_locator", {})
        if binding and locator.get("file") not in (None, binding["file"]):
            continue  # another declared file's slot: ITS record owns it
        sheet, cell = locator.get("sheet"), locator.get("cell")
        if not sheet or not cell:
            continue  # organizational rows have no writable target
        row = {"slot_id": slot_id, "sheet": sheet, "cell": cell}
        if slot.get("ref_id"):
            row["ref_id"] = slot["ref_id"]
        writable, why = _writable(slot)
        answer = prose_by_slot.get(slot_id, {})
        if slot_id not in planned_slots:
            row["decision"] = "refused_unnamed"
            row["reason"] = ("the frozen plan names no section for this "
                            "slot — T6: write-back may only touch what "
                            "the pursuit plan names")
        elif not writable:
            row["decision"] = "refused_shape"
            row["reason"] = why
        elif answer.get("status") == "drafted" and answer.get("prose"):
            row["decision"] = "written"
            row["reason"] = "drafted prose from the validated envelope"
            row["after"] = answer["prose"]
        else:
            row["decision"] = "empty_no_prose"
            row["reason"] = (f"slot status "
                            f"{answer.get('status', 'undrafted')!r} — the "
                            "cell stays honestly empty for the human")
        cells.append(row)
    workbook = load_workbook(source)
    for row in cells:
        if row["decision"] == "written":
            sheet = workbook[row["sheet"]]
            before = sheet[row["cell"]].value
            row["before"] = "" if before is None else str(before)
    facts = {
        "pursuit_id": pursuit.pursuit_id,
        "plan_sha256": envelope["plan_sha256"],
        "draft_sha256": hashlib.sha256(
            (pursuit.root / "drafts" / "draft.json").read_bytes()
        ).hexdigest(),
        "revision_n": envelope["revision_n"],
        "confirmed_by": confirmed_by,
        "at": at,
        "source_file": str(source.relative_to(pursuit.root)),
        "output_file": f"{WRITEBACK_DIR}/{source.name}",
        "cells": cells,
    }
    return facts


def preview_writeback(pursuit, *, at: str,
                      binding: dict | None = None) -> dict:
    """The S7 preview: the facts WITHOUT the write (confirmed_by is a
    placeholder the confirm step replaces — a preview confirms nothing)."""
    facts = compute_facts(pursuit, at=at, confirmed_by="(unconfirmed)",
                          binding=binding)
    return facts


def run_writeback(pursuit, log, *, at: str, confirmed_by: str,
                  binding: dict | None = None) -> dict:
    """The confirmed write: re-derives the facts server-side, copies the
    buyer file, assigns ONLY the written cells, records everything."""
    facts = compute_facts(pursuit, at=at, confirmed_by=confirmed_by,
                          binding=binding)
    source = pursuit.root / facts["source_file"]
    output = pursuit.root / facts["output_file"]
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, output)
    workbook = load_workbook(output)
    for row in facts["cells"]:
        if row["decision"] == "written":
            workbook[row["sheet"]][row["cell"]] = row["after"]
    workbook.save(output)
    facts_path = pursuit.write_artifact(
        "writeback_facts", facts,
        name=binding["facts_name"] if binding else FACTS_NAME)
    log.emit("artifact", stage="write_back", artifact={
        "kind": "writeback_facts", "path": str(facts_path),
        "revision_n": facts["revision_n"],
        "sha256": hashlib.sha256(facts_path.read_bytes()).hexdigest()})
    log.emit("artifact", stage="write_back", artifact={
        "kind": "write_back_file", "path": str(output),
        "revision_n": facts["revision_n"],
        "sha256": hashlib.sha256(output.read_bytes()).hexdigest()})
    return facts
